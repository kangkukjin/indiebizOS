"""ingest_engine — 다형 입력(텍스트/텍스트파일/CSV/PDF/엑셀/이미지) → 도메인 스키마 JSON 레코드 공용 엔진.

층 분해 (2026-08-14 설계 — "여러 형태의 입력 → 정보 추출 → 원장 적재"는 도메인마다
다시 만들 일이 아니다):
  ① 운반   = 계기 file 입력 프리미티브 + POST /launcher/upload      (표면 공용)
  ② 원문   = extract_source()   — 형식별이지 도메인별이 아님          (이 모듈)
  ③ 구조화 = extract_records()  — 제약 프롬프트 + JSON 결정론 파싱    (이 모듈)
  ④ 적재   = 도메인 op(각 원장 save) — 스키마·검증만 도메인 것
첫 소비자 = [self:health]{op:"ingest"}. 재무 어휘 등 후속 도메인은 스키마 프롬프트
하나와 op 한 줄로 이 엔진에 얹는다.

모델 선택 원칙:
  텍스트 = 경량 AI(consciousness_agent.oneshot_ai_call — 기어 리졸버 존중).
  이미지 = Gemini 비전 직접 호출 — 모달리티는 기어 무관 패스스루(model_gear _doc)이고,
           현 경량(딥시크)은 비전이 없다(2026-08-13 전환). ★모델명은 gemini-2.5-flash 명시
           (flash-latest+thinkingBudget:0=400 함정, 2026-07-22 실측).
"""
import os
import json
import re
import base64

_TEXT_EXTS = {'.txt', '.md', '.csv'}
_XLSX_EXTS = {'.xlsx', '.xls'}
_IMG_MIME = {
    '.png': 'image/png', '.jpg': 'image/jpeg', '.jpeg': 'image/jpeg',
    '.webp': 'image/webp', '.gif': 'image/gif',
}
_PDF_MIN_CHARS_PER_PAGE = 15   # 스캔(이미지) PDF 판정 (notebook §4-1 선례)
_XLSX_MAX_ROWS = 500           # 시트당 행 캡 — 초과분은 정직하게 표기
_XLSX_MAX_SHEETS = 5
_TEXT_CAP = 60_000             # 경량 모델 프롬프트 예산 방어


def _read_text_file(path: str) -> str:
    """utf-8 → cp949 → replace 인코딩 폴백 ([self:grep] 2층화 부류 — 옛 한글 파일 침묵 탈락 방지)."""
    for enc in ('utf-8', 'cp949'):
        try:
            with open(path, encoding=enc) as f:
                return f.read()
        except UnicodeDecodeError:
            continue
    with open(path, encoding='utf-8', errors='replace') as f:
        return f.read()


def _pdf_text(path: str):
    """PDF 텍스트층 추출 + 스캔 판정. returns (text, err)."""
    try:
        import fitz  # PyMuPDF
    except ImportError:
        return None, "PyMuPDF(fitz) 미설치 — PDF를 쓰려면 .venv에 pymupdf를 설치하세요."
    try:
        doc = fitz.open(path)
    except Exception as e:
        return None, f"PDF 열기 실패: {e}"
    try:
        pages = [page.get_text() for page in doc]
    finally:
        doc.close()
    text = "\n".join(pages).strip()
    if pages and len(text) < _PDF_MIN_CHARS_PER_PAGE * len(pages):
        return None, ("스캔(이미지) PDF로 보입니다(텍스트 층 없음) — "
                      "페이지를 사진으로 찍거나 스크린샷 이미지로 올려주세요.")
    return text, None


def _xlsx_text(path: str):
    """엑셀 → 시트별 TSV 텍스트. returns (text, err). 캡 초과는 정직 표기(침묵 절단 금지)."""
    try:
        import openpyxl
    except ImportError:
        return None, "openpyxl 미설치 — 엑셀을 쓰려면 .venv에 openpyxl을 설치하세요."
    try:
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    except Exception as e:
        return None, f"엑셀 열기 실패: {e}"
    parts = []
    try:
        for si, ws in enumerate(wb.worksheets):
            if si >= _XLSX_MAX_SHEETS:
                parts.append(f"[이하 시트 {len(wb.worksheets) - si}개 생략]")
                break
            rows = []
            for ri, row in enumerate(ws.iter_rows(values_only=True)):
                if ri >= _XLSX_MAX_ROWS:
                    rows.append(f"[이하 행 생략 — 시트 전체 {ws.max_row}행]")
                    break
                cells = ["" if c is None else str(c) for c in row]
                if any(c.strip() for c in cells):
                    rows.append("\t".join(cells).rstrip())
            if rows:
                parts.append(f"### 시트: {ws.title}\n" + "\n".join(rows))
    finally:
        wb.close()
    if not parts:
        return None, "엑셀에 읽을 내용이 없습니다."
    return "\n\n".join(parts), None


def extract_source(path: str = None, text: str = None) -> dict:
    """② 원문 추출 — 파일/텍스트를 모델이 먹을 수 있는 형태로.

    returns {ok: True, kind: 'text'|'image', text: str, images: list|None, label: str}
         or {ok: False, error: str}   (정직 거부 — 조용한 빈 결과 금지)
    path 와 text 를 같이 주면 text 는 사용자 메모/지시로 덧붙는다.
    """
    text = (text or "").strip()
    if not path:
        if not text:
            return {"ok": False, "error": "입력이 없습니다 — text 또는 file 을 주세요."}
        return {"ok": True, "kind": "text", "text": text[:_TEXT_CAP], "images": None, "label": "텍스트 입력"}

    if not os.path.isfile(path):
        return {"ok": False, "error": f"파일을 찾을 수 없습니다: {path}"}
    ext = os.path.splitext(path)[1].lower()
    label = os.path.basename(path)

    if ext in _IMG_MIME:
        try:
            with open(path, 'rb') as f:
                b64 = base64.b64encode(f.read()).decode('ascii')
        except OSError as e:
            return {"ok": False, "error": f"이미지 읽기 실패: {e}"}
        return {"ok": True, "kind": "image", "text": text,
                "images": [{"base64": b64, "media_type": _IMG_MIME[ext]}], "label": label}

    if ext == '.pdf':
        body, err = _pdf_text(path)
    elif ext in _XLSX_EXTS:
        body, err = _xlsx_text(path)
    elif ext in _TEXT_EXTS:
        try:
            body, err = _read_text_file(path), None
        except OSError as e:
            body, err = None, f"파일 읽기 실패: {e}"
    else:
        return {"ok": False, "error": f"지원하지 않는 형식: {ext} — 텍스트/CSV/PDF/엑셀/이미지만 받습니다."}

    if err:
        return {"ok": False, "error": err}
    if not (body or "").strip():
        return {"ok": False, "error": f"{label}: 내용이 비어 있습니다."}
    merged = (f"{body}\n\n[사용자 메모] {text}" if text else body)[:_TEXT_CAP]
    return {"ok": True, "kind": "text", "text": merged, "images": None, "label": label}


# ── ③ 구조화 추출 ──

def _strip_json(raw: str):
    """모델 출력에서 JSON 배열/객체를 결정론 추출 — 펜스·서문 관용."""
    if not raw:
        return None
    s = raw.strip()
    s = re.sub(r'^```(?:json)?\s*|\s*```$', '', s, flags=re.MULTILINE).strip()
    # 첫 [ 또는 { 부터 짝 맞는 끝까지 — ★위치 우선 (F14-4, 2026-08-20): 옛 구현은 [ 를
    # 위치 무관 우선해 {"title":…, "blocks":[…]} 객체에서 blocks 배열만 뽑았다(문서
    # "첫 [ 또는 {"와 구현의 드리프트 — table:structure 관문 이관에서 실측). 먼저
    # 나타나는 여는 괄호부터 시도하고, 파싱 실패 시 다른 쪽을 시도한다.
    candidates = [(s.find(o), o, c) for o, c in (('[', ']'), ('{', '}')) if s.find(o) >= 0]
    for i, opener, closer in sorted(candidates):
        depth = 0
        in_str = False
        esc = False
        for j in range(i, len(s)):
            ch = s[j]
            if esc:
                esc = False
                continue
            if ch == '\\':
                esc = True
            elif ch == '"':
                in_str = not in_str
            elif not in_str:
                if ch == opener:
                    depth += 1
                elif ch == closer:
                    depth -= 1
                    if depth == 0:
                        try:
                            return json.loads(s[i:j + 1])
                        except ValueError:
                            break
        # 이 opener 로 실패 → 다음 opener 시도
    return None


def _gemini_vision_json(prompt: str, images: list):
    """이미지 → Gemini REST (모달리티 패스스루). returns (raw_text, err)."""
    api_key = (os.environ.get("GEMINI_API_KEY") or "").strip()
    if not api_key:
        return None, "GEMINI_API_KEY 없음 — 이미지 추출을 하려면 .env에 키를 넣으세요."
    try:
        import requests
    except ImportError:
        return None, "requests 미설치"
    parts = [{"text": prompt}]
    for img in images:
        parts.append({"inline_data": {"mime_type": img["media_type"], "data": img["base64"]}})
    try:
        r = requests.post(
            "https://generativelanguage.googleapis.com/v1beta/models/gemini-2.5-flash:generateContent",
            headers={"x-goog-api-key": api_key, "Content-Type": "application/json"},
            json={"contents": [{"parts": parts}],
                  "generationConfig": {"temperature": 0.1, "maxOutputTokens": 4096,
                                       "thinkingConfig": {"thinkingBudget": 0}}},
            timeout=60,
        )
    except Exception as e:
        return None, f"Gemini 호출 실패: {e}"
    if r.status_code != 200:
        return None, f"Gemini {r.status_code}: {r.text[:200]}"
    try:
        return r.json()["candidates"][0]["content"]["parts"][0]["text"], None
    except (KeyError, IndexError, ValueError):
        return None, f"Gemini 응답 형식 이상: {r.text[:200]}"


def _lightweight_json(prompt: str, system_prompt: str):
    """텍스트 → 경량 AI (기어 리졸버 존중). returns (raw_text, err)."""
    try:
        from consciousness_agent import oneshot_ai_call
    except ImportError as e:
        return None, f"oneshot_ai_call 임포트 불가(백엔드 밖 실행?): {e}"
    raw = oneshot_ai_call(prompt, system_prompt=system_prompt, role="classify")
    if not raw:
        return None, "경량 AI 응답 없음"
    return raw, None


def extract_records(source: dict, schema_prompt: str, domain_label: str = "기록"):
    """③ 구조화 — 원문 + 도메인 스키마 프롬프트 → 레코드 리스트.

    source: extract_source() 결과(ok=True인 것).
    schema_prompt: 도메인이 주는 출력 계약(JSON 배열 스키마 + 규칙).
    returns (records: list, err: str|None). 파싱 실패는 모델 출력 머리를 실어 정직 반환.
    """
    system = (
        f"너는 {domain_label} 정리기다. 원문에서 저장할 기록을 추출해 JSON 배열로만 출력한다. "
        "규칙: ①원문에 없는 수치·날짜를 지어내지 말 것 ②확실치 않은 필드는 생략 "
        "③날짜는 YYYY-MM-DD ④JSON 밖에 다른 글자를 쓰지 말 것.\n\n" + schema_prompt
    )
    if source.get("kind") == "image":
        prompt = system + "\n\n[이미지에서 추출]" + (f"\n[사용자 메모] {source['text']}" if source.get("text") else "")
        raw, err = _gemini_vision_json(prompt, source["images"])
    else:
        raw, err = _lightweight_json(f"[원문]\n{source['text']}", system)
    if err:
        return [], err
    parsed = _strip_json(raw)
    if parsed is None:
        return [], f"모델 출력을 JSON으로 못 읽음: {str(raw)[:200]}"
    if isinstance(parsed, dict):
        parsed = parsed.get("records") or parsed.get("items") or [parsed]
    if not isinstance(parsed, list):
        return [], f"JSON 배열이 아님: {type(parsed).__name__}"
    return [r for r in parsed if isinstance(r, dict)], None
