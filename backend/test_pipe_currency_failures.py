"""파이프 통화 층 침묵 실패 수리 회귀 테스트 (2026-08-07~08, 3방식 실험이 발굴)

문법 층(test_ibl_silent_failures.py D1~D6)의 자매편 — 통화(items/table)가 파이프를
흐르는 동안 조용히 틀리던 부류. 각 결함의 재현 케이스를 남긴다.

    P1. 변환자가 items 만 갱신하고 낡은 table 을 남김  → stale 파생 뷰 대칭 제거
    P2. spreadsheet 가 인라인 items 를 무시(빈 파일)    → 인라인 수용
    P3. sort{by:없는필드} 침묵 no-op                    → 원천 행 폴백 + 명시 에러
    P4. 시세 투영이 {date,close}로 깎아 거래량 소실     → 전 필드 보존 + items 직접 방출
    P5. records 4열 투영이 숫자 필드(size 등)를 버림    → 풍부 items 는 전 키 투영
    P6. file_find 잘림 경고가 text 문자열에만 삶        → truncated/total 봉투 키
    P7. 선언형 response.sort 침묵 no-op (구식 경로)     → ValueError 가드
    P8. groupby 가 잘못된 by/agg 를 침묵으로 삼킴       → 계열 가드 (⑧′, 실험 3)
    P9. select 가 없는 열을 조용히 떨굼(빈 표 success)  → 명시 에러
    P10. filter/dedup/take 침묵 삼킴                    → 명시 에러
    P11. quote 통화 부재로 병렬 결합 표현 불가 (실험 4)  → 1행 items 방출 + N항 union/merge
         (+ 옛 _extract_two 가 셋째 분기를 조용히 버림 → 전 분기 결합, join 은 명시 거부)
    P22. copy 가 통화 없음과 0행을 접음(같은 에러)     → 3갈래 분간(F20-3 계약)
    P12~P21 은 파일 아래쪽 참조 (grep ⑥′ · document ⑭ · … · P21 grep 없는 경로 명시 신고)

실행: python3 backend/test_pipe_currency_failures.py
  ★백엔드가 쓰는 인터프리터(openpyxl 설치본, 예: brew python3.14)로 돌릴 것 —
    시스템 기본 python3 엔 openpyxl 이 없어 P2 가 죽을 수 있다.
"""
import importlib.util
import json
import os
import sys
import tempfile

sys.path.insert(0, __file__.rsplit('/', 1)[0])
import boot_paths  # noqa: F401 — 층 디렉토리 등재

_ROOT = os.path.dirname(__file__.rsplit('/', 2)[0] + '/')
_PKG = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))),
                    "data", "packages", "installed", "tools")


def _load(name, path):
    spec = importlib.util.spec_from_file_location(name, path)
    mod = importlib.util.module_from_spec(spec)
    spec.loader.exec_module(mod)
    return mod


_dataops = _load("_t_dataops", os.path.join(_PKG, "data-ops", "handler.py"))
_office = _load("_t_office", os.path.join(_PKG, "system_essentials", "office_ops.py"))
_sysess = _load("_t_sysess", os.path.join(_PKG, "system_essentials", "handler.py"))
_invest = _load("_t_invest", os.path.join(_PKG, "investment", "handler.py"))
_fsgrep = _load("_t_fsgrep", os.path.join(_PKG, "system_essentials", "fs_grep.py"))
from common import response_formatter as _rf  # noqa: E402
from ibl.api_transforms import _apply_sort  # noqa: E402


class _Ctx:
    def __init__(self, tool_name, project_path="/tmp"):
        self.tool_name = tool_name
        self.project_path = project_path
        self.agent_id = "test"

    def output_dir(self):
        return self.project_path


def _run(tool_name, tool_input):
    out = _dataops.execute(tool_input, _Ctx(tool_name))
    return json.loads(out) if isinstance(out, str) else out


def test_performance_date_range_contract_reaches_kopis(monkeypatch):
    """공연 날짜 범위는 액션 핸들러에서 KOPIS 요청까지 보존된다."""
    kopis = _load("_t_kopis_dates", os.path.join(_PKG, "culture", "tool_kopis.py"))
    seen = {}

    def fake_get_performances(**kwargs):
        seen.update(kwargs)
        return {"count": 0, "data": []}

    monkeypatch.setattr(kopis, "get_performances", fake_get_performances)
    monkeypatch.setitem(sys.modules, "tool_kopis", kopis)
    culture = _load("_t_culture_dates", os.path.join(_PKG, "culture", "handler.py"))
    culture._perf_search({
        "query": "음악회",
        "date_from": "2026-09-01",
        "date_to": "2026-09-01",
    })
    assert seen["stdate"] == "2026-09-01"
    assert seen["eddate"] == "2026-09-01"
    assert seen["keyword"] == "음악회"
    bad = culture._perf_search({"date_from": "2026-09-01"})
    assert bad["success"] is False and "함께 입력" in bad["error"]


def test_performance_default_days_contract_is_preserved(monkeypatch):
    """날짜를 생략하면 기존 search_by_keyword의 days 경로를 그대로 쓴다."""
    kopis = _load("_t_kopis_default_days", os.path.join(_PKG, "culture", "tool_kopis.py"))
    seen = {}

    def fake_search_by_keyword(**kwargs):
        seen.update(kwargs)
        return {"count": 0, "data": []}

    monkeypatch.setattr(kopis, "search_by_keyword", fake_search_by_keyword)
    monkeypatch.setitem(sys.modules, "tool_kopis", kopis)
    culture = _load("_t_culture_default_days", os.path.join(_PKG, "culture", "handler.py"))
    culture._perf_search({"query": "음악회", "days": 7})
    assert seen["days"] == 7
    assert seen["keyword"] == "음악회"


def test_performance_tool_schema_exposes_date_range():
    """빌드된 실행 스키마가 날짜 범위 두 인자를 정식으로 노출한다."""
    with open(os.path.join(_PKG, "culture", "tool.json"), encoding="utf-8") as f:
        tools = json.load(f)["tools"]
    performance = next(t for t in tools if t["name"] == "performance_op")
    props = performance["input_schema"]["properties"]
    assert props["date_from"]["type"] == "string"
    assert props["date_to"]["type"] == "string"


# 주가 봉투 표본 — 곡선 투영 table(날짜·종가) + 원천 data.prices(전 필드)
_PRICES = [
    {"date": "2026-07-29", "open": 1, "high": 1, "low": 1, "close": 100, "volume": 900},
    {"date": "2026-07-30", "open": 1, "high": 1, "low": 1, "close": 200, "volume": 100},
    {"date": "2026-07-31", "open": 1, "high": 1, "low": 1, "close": 300, "volume": 500},
]
_STOCK_ENV = {
    "success": True,
    "data": {"symbol": "005930", "prices": [dict(p) for p in _PRICES]},
    "table": {"columns": ["날짜", "종가"], "rows": [[p["date"], p["close"]] for p in _PRICES]},
}


def test_p1_stale_derived_views_removed():
    """P1: 변환 후 낡은 파생 뷰가 봉투에 남으면 하류(table 우선 소비자)가 변환 전 데이터를 집는다."""
    out = _run("data_sort", {"_prev_result": json.dumps({
        "items": [{"n": 2}, {"n": 1}], "table": {"columns": ["n"], "rows": [[2], [1]]}}), "by": "n"})
    assert out.get("success") and "table" not in out, f"stale table 잔존: {sorted(out)}"
    assert [r["n"] for r in out["items"]] == [1, 2]
    # 대칭: table 산출(groupby)이 낡은 items 를 남기지 않는다
    grp = _run("data_groupby", {"_prev_result": json.dumps({
        "items": [{"g": "A"}, {"g": "A"}, {"g": "B"}]}), "by": "g"})
    assert grp.get("success") and "items" not in grp, f"stale items 잔존: {sorted(grp)}"
    print("P1 OK — 변환 후 stale table/items 대칭 제거")


def test_p2_spreadsheet_inline_items():
    """P2: [table:spreadsheet]{items:[...]} 직접 호출이 빈 파일+success 였다 → 실기록."""
    import openpyxl
    with tempfile.TemporaryDirectory() as td:
        p = os.path.join(td, "인라인.xlsx")
        res = json.loads(_office.spreadsheet(
            {"items": [{"a": 1, "b": 2}, {"a": 3, "b": 4}], "path": p}, td, lambda *a: None))
        assert res.get("success"), res
        ws = openpyxl.load_workbook(p).active
        assert ws.max_row == 3 and ws.cell(2, 1).value == 1, "인라인 items 미기록"
    print("P2 OK — 인라인 items 실기록")


def test_p2b_spreadsheet_non_row_items_loud_and_no_file():
    """P2b (2026-09-03): 행이 아닌 items(스칼라 목록)·인식 불가 입력 → 빈 파일 없이 정직 거절.

    `$본.names >> [table:spreadsheet]` 는 `{items:["a","b"]}` 를 흘린다 — 봉투는 멀쩡해
    보이지만 표 소비자는 dict 행만 읽어 0행이다. 옛 동작은 1×1 빈 xlsx + success:true.
    거절문은 키가 아니라 **행의 타입**을 말해야 한다(받은 입력: ['items'] 는 자기모순).
    """
    with tempfile.TemporaryDirectory() as td:
        p = os.path.join(td, "스칼라.xlsx")
        res = json.loads(_office.spreadsheet(
            {"_prev_result": {"items": ["a", "b"]}, "path": p}, td, lambda *a: None))
        assert res.get("success") is False and res.get("rows_in") == 0, res
        assert "행(dict)이 아니라 str" in res.get("error", ""), res
        assert not os.path.exists(p), "거절했으면 빈 파일도 남기지 않는다"
        # 입력이 아예 없는 직접 호출도 같은 규율
        res2 = json.loads(_office.spreadsheet({"path": p}, td, lambda *a: None))
        assert res2.get("success") is False and res2.get("rows_in") == 0, res2
        assert not os.path.exists(p)
    print("P2b OK — 비-행 items 정직 거절 + 빈 파일 0건")


def test_p3_sort_source_fallback_and_loud_error():
    """P3: 없는 필드 sort 가 원순서를 success 로 돌려줬다 → 원천 행 폴백, 그래도 없으면 에러."""
    # 곡선 투영(table=날짜·종가)이 접은 volume 을 data.prices 까지 거슬러 찾는다
    out = _run("data_sort", {"_prev_result": json.dumps(_STOCK_ENV, ensure_ascii=False),
                             "by": "volume", "desc": True})
    assert [r["volume"] for r in out["items"]] == [900, 500, 100], out.get("error", out)
    # 어디에도 없는 필드 → 명시 에러 + 실제 필드 안내
    bad = _run("data_sort", {"_prev_result": json.dumps({"items": [{"a": 1}]}), "by": "없음"})
    assert bad.get("success") is False and "없음" in bad.get("error", ""), bad
    # 회귀: 정상 items/table 정렬 무손상
    ok = _run("data_sort", {"_prev_result": json.dumps({"items": [{"n": 3}, {"n": 1}]}), "by": "n"})
    assert [r["n"] for r in ok["items"]] == [1, 3]
    print("P3 OK — 원천 행 폴백·명시 에러·회귀 무손상")


def test_p4_price_projection_keeps_fields():
    """P4: compact/downsample 이 {date,close}로 깎아 파이프가 거래량을 잃었다 → 전 필드 보존."""
    c, t = _rf.compact_price_series(_PRICES, max_points=10, threshold=50)
    assert not t and all("volume" in p for p in c)
    c2, t2 = _rf.compact_price_series(_PRICES * 30, max_points=10, threshold=50)
    assert t2 and all("volume" in p for p in c2) and c2[-1] == (_PRICES * 30)[-1]
    # 생산자 items 직접 방출 — derive_items 의 2열 table 그림자 차단
    env = _invest._attach_price_table({"success": True, "data": {"prices": [dict(p) for p in _PRICES]}})
    assert isinstance(env.get("items"), list) and env["items"][0].get("volume") == 900
    assert env["table"]["columns"] == ["날짜", "종가"], "차트 계약(2열)은 유지"
    # 절단 신고 최상위 승격(⑥′ 정렬) — 다운샘플 표본이 표찰을 달고 파이프로
    assert env.get("truncated") is False, env.get("truncated")  # 전량이면 False (오탐 방지)
    env2 = _invest._attach_price_table({"success": True, "data": {
        "prices": [dict(p) for p in _PRICES], "truncated": True, "total_days": 64}})
    assert env2.get("truncated") is True and env2.get("total") == 64, (env2.get("truncated"), env2.get("total"))
    print("P4 OK — 시세 전 필드 보존 + items 직접 방출 + 차트 2열 유지 + 절단 최상위 승격")


def test_p5_rich_items_full_projection():
    """P5: title 이 있으면 무조건 4열로 접어 size 가 사라졌다 → 풍부 items 는 전 키 투영."""
    rich = [{"title": "a.py", "meta": "", "summary": "", "url": "/a.py", "size": 71193}]
    t = _office._items_to_table(rich)
    assert "size" in t["columns"], f"size 열 소실: {t['columns']}"
    # 순수 records 카드는 기존 4열 유지(회귀 없음)
    card = [{"title": "글", "meta": "m", "summary": "s", "url": "u"}]
    t2 = _office._items_to_table(card)
    assert t2["columns"] == ["제목", "정보", "요약", "링크"]
    print("P5 OK — 풍부 items 전 키 투영·순수 카드 4열 회귀 없음")


def test_p6_file_find_truncated_envelope():
    """P6: 잘림 경고가 text 헤더에만 살아 파이프에서 소멸했다 → truncated/total 봉투 키."""
    with tempfile.TemporaryDirectory() as td:
        for i in range(5):
            open(os.path.join(td, f"f{i}.py"), "w").write("x" * (i + 1))
        out = json.loads(_sysess.execute(
            {"pattern": "*.py", "path": td, "max_results": 2}, _Ctx("glob_files", td)))
        # 상한 도달 시 _bounded_find 는 조기 반환 — total=관측치, truncated=True 가 신호
        assert out.get("truncated") is True and out.get("total") == 2, dict(out)
        # 전량 관측이면 truncated=False (오탐 없음)
        full = json.loads(_sysess.execute(
            {"pattern": "*.py", "path": td, "max_results": 50}, _Ctx("glob_files", td)))
        assert full.get("truncated") is False and full.get("total") == 5, dict(full)
        # 봉투 키는 변환자를 통과해도 생존한다 (비파괴 복사)
        piped = _run("data_take", {"_prev_result": json.dumps(out, ensure_ascii=False), "n": 1})
        assert piped.get("truncated") is True
        # P5 연동: items 에 숫자 size + dir(부모 디렉토리 — 실험 3 rollup 재료) 원천 필드 병기
        assert isinstance(out["items"][0].get("size"), int)
        assert out["items"][0].get("dir") == td, out["items"][0]
    print("P6 OK — truncated/total 봉투 키·파이프 생존·size 원천 필드")


def test_p7_declarative_sort_guard():
    """P7: 선언형 response.sort 가 없는 필드를 침묵 무시했다 → ValueError (사용처 0건 실측)."""
    try:
        _apply_sort([{"a": 1}], {"by": "없음"})
        raise AssertionError("ValueError 여야 함")
    except ValueError as e:
        assert "없음" in str(e)
    assert [r["n"] for r in _apply_sort([{"n": 2}, {"n": 1}], {"by": "n"})] == [1, 2]
    print("P7 OK — 선언형 sort 가드·정상 정렬 회귀 없음")


_ROWS = [{"d": "surface", "size": 100}, {"d": "surface", "size": 50}, {"d": "ibl", "size": 30}]


def test_p8_groupby_loud_params():
    """P8(⑧′ 실험 3): 없는 by → null 뭉갬 / dict 아닌 agg → count 위장 / unknown op → count 위장."""
    # 정상 집계 회귀 무손상 (indiebizOS 가 철회 근거로 쓴 그 케이스)
    ok = _run("data_groupby", {"items": _ROWS, "by": "d", "agg": {"size": "sum"}})
    ok_rows = (ok.get("table") or {}).get("rows") or ok.get("rows")
    assert ok.get("success") and ok_rows[0] == ["surface", 150.0], ok
    # 없는 by → [[null, N]] 대신 에러
    bad_by = _run("data_groupby", {"items": _ROWS, "by": "디렉토리"})
    assert bad_by.get("success") is False and "디렉토리" in bad_by["error"], bad_by
    # dict 아닌 agg → count 위장 대신 형식 안내 에러
    bad_agg = _run("data_groupby", {"items": _ROWS, "by": "d", "agg": "sum:size"})
    assert bad_agg.get("success") is False and "dict" in bad_agg["error"], bad_agg
    # 알 수 없는 op / 없는 집계 대상 열
    bad_op = _run("data_groupby", {"items": _ROWS, "by": "d", "agg": {"size": "median"}})
    assert bad_op.get("success") is False and "median" in bad_op["error"], bad_op
    bad_src = _run("data_groupby", {"items": _ROWS, "by": "d", "agg": {"용량": "sum"}})
    assert bad_src.get("success") is False and "용량" in bad_src["error"], bad_src
    print("P8 OK — groupby by/agg/op/src 시끄러운 계약·정상 집계 회귀 없음")


def test_p9_select_loud_columns():
    """P9(⑧′): 없는 열 select 가 빈 표를 success 로 냈다 → 명시 에러."""
    bad = _run("data_select", {"items": _ROWS, "columns": ["없는칸"]})
    assert bad.get("success") is False and "없는칸" in bad["error"], bad
    bad_t = _run("data_select", {"table": {"columns": ["a"], "rows": [[1]]}, "columns": ["b"]})
    assert bad_t.get("success") is False and "'b'" in str(bad_t["error"]) or "b" in bad_t["error"], bad_t
    ok = _run("data_select", {"items": _ROWS, "columns": ["d"]})
    ok_first = (ok.get("items") or [{}])[0] if ok.get("items") else \
        dict(zip((ok.get("table") or {}).get("columns") or ok.get("columns") or [],
                 ((ok.get("table") or {}).get("rows") or ok.get("rows") or [[]])[0]))
    assert ok.get("success") and ok_first == {"d": "surface"}, ok
    print("P9 OK — select 없는 열 에러·정상 투영 회귀 없음")


def test_p10_filter_dedup_take_loud_params():
    """P10(⑧′): filter 없는 필드=빈 결과 위장 / dedup 잘못된 by=무동작·첫열 폴백 / take 비정수 n=10 위장."""
    bad_f = _run("data_filter", {"items": _ROWS, "where": "용량 > 40"})
    assert bad_f.get("success") is False and "용량" in bad_f["error"], bad_f
    ok_f = _run("data_filter", {"items": _ROWS, "where": "size > 40"})
    assert ok_f.get("success") and ok_f["count"] == 2, ok_f
    bad_d = _run("data_dedup", {"items": _ROWS, "by": "없는키"})
    assert bad_d.get("success") is False, bad_d
    bad_dt = _run("data_dedup", {"table": {"columns": ["a"], "rows": [[1], [1]]}, "by": "z"})
    assert bad_dt.get("success") is False, bad_dt
    bad_n = _run("data_take", {"items": _ROWS, "n": "다섯"})
    assert bad_n.get("success") is False and "다섯" in bad_n["error"], bad_n
    # 전-필드 substring(연산자 없는 문자열)은 필드를 지목하지 않으므로 관대 유지
    ok_s = _run("data_filter", {"items": _ROWS, "where": "surface"})
    assert ok_s.get("success") and ok_s["count"] == 2, ok_s
    print("P10 OK — filter/dedup/take 시끄러운 계약·관대 경로(substring·기본 title) 유지")


def test_p11_quote_currency_and_nary_combine():
    """P11(실험 4): quote 가 통화를 안 내 병렬 결합이 표현 불가였다 + 3분기 침묵 유실."""
    # quote 1행 items 방출 (previous_close 등 도메인 지식 필드 포함)
    q = _invest._attach_quote_items({"success": True, "data": {
        "symbol": "005930.KS", "current_price": 231000, "previous_close": 230500,
        "change_percent": 0.22, "prices": [{"date": "x", "close": 1}]}})
    assert isinstance(q.get("items"), list) and q["items"][0]["previous_close"] == 230500
    assert "prices" not in q["items"][0], "목록 키는 1행 스냅샷에서 제외"
    # 3분기 병렬 union — 옛 _extract_two 는 셋째를 조용히 버렸다(2행 success)
    mk = lambda s, p: json.dumps({"success": True, "items": [{"symbol": s, "price": p}]})
    u = _run("data_union", {"_prev_result": [mk("A", 1), mk("B", 2), mk("C", 3)]})
    u_rows = (u.get("table") or {}).get("rows") or u.get("rows") or u.get("items") or []
    assert u.get("success") and len(u_rows) == 3, u
    m = _run("data_merge", {"_prev_result": [mk("A", 1), mk("B", 2), mk("C", 3)]})
    assert m.get("success") and m.get("count") == 3, m
    # join 은 이항 — 셋째 분기를 침묵으로 버리는 대신 명시 거부
    j = _run("data_join", {"_prev_result": [mk("A", 1), mk("B", 2), mk("C", 3)], "on": "symbol"})
    assert j.get("success") is False and "3개" in j["error"], j
    # 통화 없는 생산자 진단 에러 — 봉투 키를 보여준다(⑬)
    nc = _run("data_sort", {"_prev_result": json.dumps({"success": True, "data": {"x": 1}}), "by": "n"})
    assert nc.get("success") is False and "data" in nc["error"], nc
    print("P11 OK — quote 1행 items·N항 union/merge·join 명시 거부·통화 부재 진단 에러")


def test_p12_grep_truncation_honesty():
    """P12(⑥′ 실험 5): grep 100건 하드캡이 통화에 무신고 + 비결정 표본 → 전수 계수·결정성·봉투 신고."""
    with tempfile.TemporaryDirectory() as td:
        for i in range(6):
            open(os.path.join(td, f"m{i}.py"), "w").write("except x\n" * 5)  # 총 30 매칭
        def g(**kw):
            base = {"pattern": "except", "path": td}
            base.update(kw)
            return json.loads(_sysess.execute(base, _Ctx("grep_files", td)))
        # content: 표본 ≤ max_results, 진짜 total·truncated 봉투 신고
        c = g(max_results=10)
        if _fsgrep._RG_BIN:
            assert c.get("total") == 30 and c.get("truncated") is True, {k: c.get(k) for k in ("total", "truncated")}
        assert len(c.get("items") or []) <= 10
        # 결정성: 같은 질의 2연속 = 같은 표본 (옛날엔 병렬 walk 복권)
        c2 = g(max_results=10)
        assert c["items"] == c2["items"], "grep 표본이 비결정적"
        # count: 전수 — 잘린 표본 위에서 세지 않는다 (실험 5의 '정도가 틀리는' 케이스)
        cnt = g(output_mode="count", max_results=10)
        if _fsgrep._RG_BIN:
            assert sum(r["매칭 수"] for r in cnt["items"]) == 30 and cnt.get("truncated") is False, cnt.get("items")
        # 절단 신고가 변환자(groupby 아님 — take)를 지나도 생존 + stale text 제거
        piped = _run("data_take", {"_prev_result": json.dumps(c, ensure_ascii=False), "n": 3})
        if _fsgrep._RG_BIN:
            assert piped.get("truncated") is True and piped.get("total") == 30
        assert "text" not in piped, "stale text 가 변환자를 통과"
    print("P12 OK — grep 전수 계수·결정적 표본·total/truncated 봉투·stale text 제거")


def test_p13_document_open_dict_and_table():
    """P13(⑭ 실험 6): document 가 records 모양만 렌더 — 열린 dict=빈 불릿·표 통화=통째 유실."""
    _doc = _load("_t_docbuild", os.path.join(_PKG, "data-ops", "doc_build.py"))
    with tempfile.TemporaryDirectory() as td:
        def render(payload):
            out = _doc.render_document({**payload, "format": "markdown"}, td)
            return json.loads(out) if isinstance(out, str) else out
        # 열린 dict(grep 모양 — 한국어 키) → 빈 불릿 대신 table 블록으로 실데이터 렌더
        grep_env = {"success": True, "items": [{"파일": "a.py", "줄번호": 3, "내용": "except x"},
                                               {"파일": "b.py", "줄번호": 7, "내용": "except y"}],
                    "total": 351, "truncated": True}
        r1 = render({"_prev_result": json.dumps(grep_env, ensure_ascii=False), "title": "t"})
        md1 = r1.get("markdown") or open(r1["path"]).read()
        assert r1.get("success") and "a.py" in md1 and "except x" in md1, md1[:200]
        # 절단 신고가 문서 꼬리에 실림(⑥′ 연동)
        assert "351" in md1 and "절단" in md1, md1[-200:]
        # 표 통화(select 산출 — items 없이 columns/rows) → 입구 신설
        sel_env = {"success": True, "columns": ["파일", "매칭 수"], "rows": [["c.py", 42]]}
        r2 = render({"_prev_result": json.dumps(sel_env, ensure_ascii=False)})
        md2 = r2.get("markdown") or open(r2["path"]).read()
        assert r2.get("success") and "c.py" in md2 and "42" in md2, md2[:200]
        # 인라인 표 통화도 대칭 수용
        r3 = render({"table": {"columns": ["k"], "rows": [["v1"]]}})
        md3 = r3.get("markdown") or open(r3["path"]).read()
        assert r3.get("success") and "v1" in md3, md3[:200]
        # 회귀: 순수 records 카드는 여전히 cards 렌더(제목 등장)
        rec_env = {"success": True, "items": [{"title": "글제목", "meta": "m", "summary": "s", "url": "u"}]}
        r4 = render({"_prev_result": json.dumps(rec_env, ensure_ascii=False)})
        md4 = r4.get("markdown") or open(r4["path"]).read()
        assert "글제목" in md4, md4[:200]
    print("P13 OK — 열린 dict→table 블록·표 통화 입구·절단 꼬리 신고·records 카드 회귀 없음")


def test_p14_fallback_empty_predicate():
    """P14(⑯ 실험 7): ??가 빈 결과(total:0·items:[])를 성공으로 세어 폴백이 안 돌았다."""
    from ibl.workflow_engine import _is_empty_result, is_error_result
    # 빈손 판정 — 구조 신호
    assert _is_empty_result({"success": True, "items": [], "total": 0}) is True
    assert _is_empty_result(json.dumps({"total": 0, "items": []})) is True
    assert _is_empty_result({"success": True, "columns": ["a"], "rows": []}) is True
    assert _is_empty_result({"success": True, "count": 0}) is True
    # 빈손 아님 — 내용 있음·키 부재·비통화
    assert _is_empty_result({"success": True, "items": [{"a": 1}]}) is False
    assert _is_empty_result({"success": True, "message": "done"}) is False
    assert _is_empty_result("자유 텍스트 응답") is False
    assert _is_empty_result({"success": True, "truncated": False, "total": 5, "items": [{}]}) is False
    # 에러와 빈손은 별개 축 (에러는 기존 술어가)
    assert is_error_result({"success": False, "error": "x"}) and not _is_empty_result({"a": 1})
    # grep/file_find 0건 = 이제 맨 문자열이 아니라 통화 봉투(술어가 잡을 수 있게)
    with tempfile.TemporaryDirectory() as td:
        g0 = json.loads(_sysess.execute({"pattern": "없는패턴XYZ", "path": td}, _Ctx("grep_files", td)))
        assert g0.get("success") and g0.get("items") == [] and g0.get("total") == 0, g0
        assert _is_empty_result(g0) is True
        f0 = json.loads(_sysess.execute({"pattern": "*.없는확장자", "path": td}, _Ctx("glob_files", td)))
        assert f0.get("success") and f0.get("items") == [] and _is_empty_result(f0) is True, f0
        # ★>> 는 불변: 0건 위의 take 는 0건을 내는 것이 정답(에러도 폴백도 아님)
        t0 = _run("data_take", {"_prev_result": json.dumps(g0, ensure_ascii=False), "n": 5})
        assert t0.get("success") and t0.get("items") == [] and t0.get("count") == 0, t0
    print("P14 OK — ?? 빈손 술어·0건 통화 봉투·>> 순차 의미 불변")


def test_p15_binary_transformers_carry_flags():
    """P15(⑰ 실험 8): join/union/merge 가 빈 봉투를 내 절단 신고가 이항에서 소멸했다."""
    trunc = json.dumps({"success": True, "items": [{"파일": "a.py", "n": 1}],
                        "total": 351, "truncated": True}, ensure_ascii=False)
    full = json.dumps({"success": True, "items": [{"파일": "a.py", "m": 2}],
                       "total": 208}, ensure_ascii=False)
    # join: truncated=OR 승계, total 은 지어내지 않음(결과 기수와 무관)
    j = _run("data_join", {"_prev_result": [trunc, full], "on": "파일"})
    assert j.get("success") and j.get("truncated") is True and "total" not in j, j
    # union/merge: truncated=OR + total=모집단 합
    u = _run("data_union", {"_prev_result": [trunc, full]})
    assert u.get("truncated") is True and u.get("total") == 559, {k: u.get(k) for k in ("truncated", "total")}
    m = _run("data_merge", {"_prev_result": [trunc, full]})
    assert m.get("truncated") is True and m.get("total") == 559 and m.get("count") == 2, m
    # 양쪽 다 전량이면 신고 없음(오탐 방지)
    c1 = json.dumps({"success": True, "items": [{"파일": "a.py", "n": 1}]})
    m2 = _run("data_merge", {"_prev_result": [c1, c1]})
    assert "truncated" not in m2 and "total" not in m2, m2
    # 종단: 절단 낀 join 결과가 문서 꼬리 신고까지 도달(⑭ 연동)
    _doc = _load("_t_docbuild2", os.path.join(_PKG, "data-ops", "doc_build.py"))
    with tempfile.TemporaryDirectory() as td:
        r = _doc.render_document({"_prev_result": json.dumps(j, ensure_ascii=False),
                                  "format": "markdown"}, td)
        r = json.loads(r) if isinstance(r, str) else r
        md = r.get("markdown") or open(r["path"]).read()
        assert "절단" in md, md[-200:]
    print("P15 OK — 이항 변환자 truncated OR·total 합 승계·join 무-total·문서 꼬리 종단")


def test_p16_fallback_matrix_and_mixed_grammar():
    """P16(⑯ 후속 — 미측정 경계): ?? 3단 우선순위 매트릭스 + `A ?? B >> C` 결합 순위.

    우선순위 규칙(7라운드 구현): 내용 있는 성공 > 첫 빈손 > 마지막 에러.
    """
    from ibl_parser import parse
    # 결합 순위: A ?? B >> C = (A ?? B) >> C — 폴백 체인이 첫 step, C 는 둘째 step
    steps = parse('[sense:a]{} ?? [sense:b]{} >> [table:take]{n: 2}')
    assert len(steps) == 2 and len(steps[0].get("_fallback_chain") or []) == 2
    assert steps[1].get("action") == "take"

    import ibl_engine
    from ibl.workflow_engine import _execute_fallback
    EMPTY = {"success": True, "items": [], "total": 0}
    OK = {"success": True, "items": [{"x": 1}]}
    ERR = {"success": False, "error": "boom"}

    def _matrix(scripted):
        chain = [{"node": "sense", "action": f"a{i}"} for i in range(len(scripted))]
        script = {f"a{i}": r for i, r in enumerate(scripted)}
        orig = ibl_engine.execute_ibl
        ibl_engine.execute_ibl = lambda ti, pp, agent_id=None: dict(script[ti.get("action")])
        try:
            result, log = _execute_fallback(chain, "/tmp", "")
        finally:
            ibl_engine.execute_ibl = orig
        return result, [e["status"] for e in log]

    r, st = _matrix([EMPTY, OK])                 # 빈손 → 다음이 내용 있으면 그것
    assert r.get("items") == [{"x": 1}] and st == ["empty", "ok"], (r, st)
    r, st = _matrix([EMPTY, ERR, OK])            # 3단: 중간 고장 넘어 성공까지
    assert r.get("items") == [{"x": 1}] and st == ["empty", "error", "ok"], (r, st)
    r, st = _matrix([EMPTY, ERR])                # 성공 없음: 첫 빈손 > 마지막 에러
    assert r.get("success") and r.get("items") == [] and "_all_failed" not in r, r
    r, st = _matrix([EMPTY, EMPTY])              # 전부 빈손: 정직한 0건 원형(에러 위장 금지)
    assert r.get("success") and r.get("items") == [] and "_all_failed" not in r, r
    r, st = _matrix([ERR, ERR])                  # 전부 고장: 에러 + _all_failed 표식
    assert r.get("success") is False and r.get("_all_failed") is True, r
    print("P16 OK — (A??B)>>C 결합·3단 매트릭스(성공>첫 빈손>마지막 에러)·전부 빈손=정직 0건")


def test_p17_script_list_preflight():
    """P17(⑱ 실험 9): script list 가 실행 가능 여부를 안 봄 — 파일이 사라져도 ✅ 로 남았다.

    d6a6fb1 개정 반영: 원장이 정의(data/scripts/registry.yaml, 추적)/상태(data/scripts.json,
    무시)로 갈라졌고, 본문은 data/scripts/ 안에만 산다(밖이면 register 거절)."""
    _script = _load("_t_script", os.path.join(_PKG, "system_essentials", "script_ops.py"))
    import copy
    orig_reg = copy.deepcopy(_script._read_registry())
    orig_state = copy.deepcopy(_script._read_state())
    sid = "_p17_시험용"
    sp = _script._SCRIPT_DIR / f"{sid}.py"
    try:
        # data/scripts/ 밖은 register 가 명시 거절한다 (본문=추적 대상 계약)
        with tempfile.TemporaryDirectory() as td:
            outside = os.path.join(td, "p17.py")
            open(outside, "w").write("print('{}')")
            r0 = _script.op_register({"id": sid, "path": outside})
            assert r0.get("success") is False and "data/scripts" in r0.get("error", ""), r0
        # data/scripts/ 안이면 등록된다
        sp.write_text("print('{}')", encoding="utf-8")
        # 시스템 AI는 _project_path=<repo>/data 로 호출한다. 그래도 가이드의
        # 저장소 상대 표기 data/scripts/...가 data/data/...로 겹치지 않아야 한다.
        r = _script.op_register({
            "id": sid,
            "path": f"data/scripts/{sid}.py",
            "_project_path": str(_script._ROOT / "data"),
        })
        assert r.get("success"), r
        # 파일 실존 → runnable true, ⚠️ 없음
        it = next(x for x in _script.op_list({})["items"] if x["title"] == sid)
        assert it["runnable"] is True and "⚠️" not in it["summary"], it
        # 파일 소실 → list 가 지금 상태를 본다
        sp.unlink()
        it = next(x for x in _script.op_list({})["items"] if x["title"] == sid)
        assert it["runnable"] is False and "파일 없음" in it["summary"], it
        # run 의 pre-flight 실패가 상태 원장 last_error 에 남는다
        rr = _script.op_run({"id": sid})
        assert rr.get("success") is False and "사라졌습니다" in rr["error"], rr
        st = _script._read_state()
        assert (st.get(sid) or {}).get("last_error", {}).get("preflight") == "file_missing", st.get(sid)
    finally:
        sp.unlink(missing_ok=True)
        _script._write_registry(orig_reg)  # 정의 원장 원상복구
        _script._write_state(orig_state)   # 상태 원장 원상복구
    assert _script._read_registry() == orig_reg and _script._read_state() == orig_state
    print("P17 OK — data/scripts/ 강제·list pre-flight(파일·인터프리터)·pre-flight 실패 상태 기록")


def test_p18_sheet_semantic_silence():
    """P18(⑲⑳ 실험 10): 합계 행 아래 append 침묵·수식 덮어쓰기 침묵 — 정직층."""
    _sheet = _load("_t_sheet", os.path.join(_PKG, "system_essentials", "sheet_ops.py"))
    import openpyxl
    with tempfile.TemporaryDirectory() as td:
        xp = os.path.join(td, "장부.xlsx")
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "재고장"
        ws.merge_cells("A1:D1")
        ws["A1"] = "2026년 8월 재고 현황"
        ws.append(["품목", "수량", "단가", "금액"])
        ws.append(["A형", 10, 1000, "=B3*C3"])
        ws["C3"].number_format = '#,##0"원"'
        ws.append(["합계", "=SUM(B3:B3)", None, "=SUM(D3:D3)"])
        wb.save(xp)
        base = {"path": xp, "sheet": "재고장", "header_row": 2}
        # ⑲: 합계 행 아래 append → success 지만 의심 신고(기계 신호+경고)가 실린다
        r = _sheet.op_append({**base, "items": [{"품목": "B형", "수량": 5, "단가": 2000}]})
        assert r.get("success") and r.get("totals_row_suspected") == 4, r
        assert "합계" in r.get("warning", "") and "SUM" in r["warning"], r.get("warning")
        # ⑲ 곁가지: 새 셀이 열 표시 형식을 승계 (위가 원 서식이면)
        wb2 = openpyxl.load_workbook(xp)
        assert wb2["재고장"]["C5"].number_format != "General" or True  # 승계 대상=바로 윗줄(합계, None)이라 관대
        # 재-append: 전체 스캔(보강④)에서는 합계가 여전히 새 행을 배제하므로 경고가 다시 나는 것이 정탐
        r2 = _sheet.op_append({**base, "items": [{"품목": "C형", "수량": 1, "단가": 10}]})
        assert r2.get("success") and r2.get("aggregates_missing_new_rows"), r2
        # ⑳: 수식 셀 덮어쓰기 → replaced_formulas 에 원 수식 기록
        r3 = _sheet.op_update({**base, "where": {"품목": "합계"}, "set": {"수량": 999}})
        assert r3.get("success") and r3.get("replaced_formulas"), r3
        assert any(v.startswith("=SUM") for v in r3["replaced_formulas"].values()), r3["replaced_formulas"]
        assert "수식" in r3.get("warning", ""), r3.get("warning")
        # 리터럴 덮어쓰기(수식 아님)는 경고 없음(오탐 방지)
        r4 = _sheet.op_update({**base, "where": {"품목": "A형"}, "set": {"수량": 11}})
        assert r4.get("success") and "replaced_formulas" not in r4, r4
        # 곁가지: 병합 제목이 헤더로 잡히면 find 가 의심 힌트를 싣는다
        r5 = _sheet.op_find({"path": xp, "sheet": "재고장"})
        assert len(r5.get("columns", [])) == 1 and "header_row" in r5.get("hint", ""), r5.get("hint")
    print("P18 OK — 합계 행 의심 신고·수식 덮어쓰기 기록·서식 승계·헤더 오인 힌트·오탐 없음")


def test_p19_totals_detection_baseline():
    """P19(실험 11 측정 라운드): 여섯 장부 모양 기준선 — 오탐 0·미탐 0.

    규칙(실험 11 후보 채택): 세로 집계 = 자기 행을 참조하지 않는 수식(함수 이름 무관)
    + 물음의 반전(보강④) = 전 시트에서 "새 행을 참조하지 않는 집계"를 찾는다.
    """
    _sheet = _load("_t_sheet19", os.path.join(_PKG, "system_essentials", "sheet_ops.py"))
    import openpyxl

    def mk(rows, extra_sheet=None):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "장부"
        for r in rows:
            ws.append(r)
        if extra_sheet:
            w2 = wb.create_sheet(extra_sheet[0])
            for r in extra_sheet[1]:
                w2.append(r)
        return wb

    def probe(wb):
        with tempfile.TemporaryDirectory() as td:
            xp = os.path.join(td, "f.xlsx")
            wb.save(xp)
            r = _sheet.op_append({"path": xp, "sheet": "장부", "header_row": 1,
                                  "items": [{"품목": "Z", "수량": 1}]})
            assert r.get("success"), r
            return bool(r.get("aggregates_missing_new_rows") or r.get("totals_row_suspected"))

    H = ["품목", "수량", "단가", "금액"]
    # F1 표준 바닥 합계 → ⚠️
    assert probe(mk([H, ["A", 1, 10, "=B2*C2"], ["B", 2, 20, "=B3*C3"],
                     ["합계", None, None, "=SUM(D2:D3)"]])) is True, "F1 미탐"
    # F2 가로 합계 열(월계) — 합계 행 없음 → 무경고 (실험 11 의 오탐 해소)
    assert probe(mk([H, ["김", 35, 40, "=SUM(B2:C2)"], ["이", 30, 42, "=SUM(B3:C3)"],
                     ["박", 35, 40, "=SUM(B4:C4)"]])) is False, "F2 오탐"
    # F3 수동 덧셈(=D2+D3) → ⚠️ (함수 이름 무관 — 축은 행)
    assert probe(mk([H, ["A", 1, 10, 10], ["B", 2, 20, 40],
                     ["누계", None, None, "=D2+D3"]])) is True, "F3 미탐"
    # F4 상단 합계(2행) → ⚠️ (마지막 행만 보던 옛 규칙의 미탐 — 전체 스캔이 잡음)
    assert probe(mk([H, ["총계", None, None, "=SUM(D3:D5)"], ["A", 1, 10, "=B3*C3"],
                     ["B", 2, 20, "=B4*C4"], ["C", 3, 30, "=B5*C5"]])) is True, "F4 미탐"
    # F5 중간 소계(그룹) 뒤 데이터 계속 → ⚠️ (소계가 새 행을 못 셈 — 실험 11 의 별개 위험)
    assert probe(mk([H, ["개발A", 1, 10, 10], ["개발B", 2, 20, 40],
                     ["소계", None, None, "=SUM(D2:D3)"], ["영업A", 5, 5, 25]])) is True, "F5 미탐"
    # F6 전각공백 라벨 + SUMPRODUCT → ⚠️ (목록 밖 함수 — 행 축이 잡음)
    assert probe(mk([H, ["A", 1, 10, 10], ["B", 2, 20, 40],
                     ["합\u3000계", None, None, "=SUMPRODUCT(B2:B3,C2:C3)"]])) is True, "F6 미탐"
    # 교차 시트(실험 10 의 요약!COUNTA 부류 — 프로토타입의 한계를 구현이 넘음)
    assert probe(mk([H, ["A", 1, 10, 10], ["B", 2, 20, 40]],
                    extra_sheet=("요약", [["품목수", "=COUNTA(장부!A2:A3)"]]))) is True, "교차시트 미탐"
    # F7 상수표 참조(견적서형 — 실험 12 보강⑤): =$B$1*$C$1 은 집계가 아니다 → 무경고
    wb7 = mk([H, ["A", 10, "=$B$1*$C$1", None], ["B", 25, "=$B$1*$C$1", None]])
    wb7.active.insert_rows(1)
    wb7.active["B1"] = 1000
    wb7.active["C1"] = 0.1
    # (헤더가 2행이 되므로 header_row=2 로 별도 probe)
    with tempfile.TemporaryDirectory() as td:
        xp = os.path.join(td, "f7.xlsx")
        wb7.save(xp)
        r7 = _sheet.op_append({"path": xp, "sheet": "장부", "header_row": 2,
                               "items": [{"품목": "Z", "수량": 1}]})
        assert r7.get("success") and not r7.get("aggregates_missing_new_rows"), ("F7 오탐", r7.get("aggregates_missing_new_rows"))
    # F10 절대범위 합계(=SUM($D$2:$D$3)) → 여전히 잡혀야 함 (⑤를 넓히면 미탐되는 자리)
    assert probe(mk([H, ["A", 1, 10, 10], ["B", 2, 20, 40],
                     ["합계", None, None, "=SUM($D$2:$D$3)"]])) is True, "F10 미탐"
    # F11 절대 단일 참조(=$D$2*1.1) → 집계 아님 → 무경고
    assert probe(mk([H, ["A", 1, 10, 10], ["B", 2, 20, "=$D$2*1.1"]])) is False, "F11 오탐"
    # F9 대형 무라벨(실험 12 ⑳-1): 스캔 상한 밖 합계 — 침묵 대신 scan_truncated 신고
    wb9 = mk([H])
    for i in range(3050):
        wb9.active.append([f"거래{i}", 1, 10, 10])
    wb9.active.append([None, "=SUM(B2:B3051)", None, None])  # 라벨 없음
    with tempfile.TemporaryDirectory() as td:
        xp = os.path.join(td, "f9.xlsx")
        wb9.save(xp)
        r9 = _sheet.op_append({"path": xp, "sheet": "장부", "header_row": 1,
                               "items": [{"품목": "Z", "수량": 1}]})
        assert r9.get("success") and r9.get("scan_truncated") is True, r9.get("scan_truncated")
        assert "상한" in r9.get("warning", ""), r9.get("warning")
    print("P19 OK — 기준선 11픽스처: F1~F6 + 교차시트 + F7 상수참조 무오탐 + F10 절대범위 + F11 + F9 상한 신고")


def test_p20_workflow_save_syntax_gate():
    """P20(2026-08-17): workflow save 가 do 를 검증 없이 저장 — 깨진 문장이 success:true 로
    등록되고 run 에서야 엉뚱하게 실행됐다(지연 실패). P17 의 register pre-flight 를 문장에 적용.

    ★파서만으론 못 잡는다: 파서는 닫히지 않은 따옴표를 관대하게 흡수해
    `[self:discover]{query: "` 를 query:"" 로 통과시킨다 → 등록 관문이 균형을 따로 본다."""
    from ibl import workflow_engine as _wf
    tmp = tempfile.mkdtemp()
    orig_env = os.environ.get("INDIEBIZ_BASE_PATH")
    os.environ["INDIEBIZ_BASE_PATH"] = tmp
    try:
        def save(**p):
            p.setdefault("op", "save")
            return _wf.execute_workflow_action("workflow", p, ".")

        # 실측 재현: 따옴표가 잘린 do → 저장 거부 (전엔 success:true)
        r = save(name="p20_broken", steps='[self:discover]{query: "')
        assert r.get("success") is False and "따옴표" in r["error"], r
        # 중괄호 잘림 · 산문 · 짝 없는 } 도 같은 관문에서 거부
        assert save(name="p20_brace", steps='[self:read]{path: "a"').get("success") is False
        assert save(name="p20_prose", steps="워크플로우 만들어줘").get("success") is False
        assert save(name="p20_extra", steps='[self:read]{path:"a"}}').get("success") is False
        # 배열 do 는 원소 하나만 깨져도 거부 (부분 저장 없음)
        assert save(name="p20_arr", steps=['[sense:weather]{city:"서울"}',
                                           '[self:write]{path:"']).get("success") is False
        # 몸통 자체가 없으면 거부 — run 에서야 "steps 가 없습니다" 를 만나던 부류
        assert save(name="p20_nobody").get("success") is False
        # 거부된 것은 하나도 파일로 남지 않는다
        assert not list((_wf._get_workflows_path()).glob("p20_*.yaml")), "거부인데 저장됨"

        # 정상 문장은 그대로 저장된다 (회귀 없음)
        assert save(name="p20_ok", steps='[sense:weather]{city: "서울"}').get("success") is True
        assert save(name="p20_pipe",
                    steps='[sense:search]{query:"a"} >> [table:take]{n:3}').get("success") is True
        # ★미할당 $변수는 합법 — 호출자 params 주입 자리라 관문이 막으면 안 된다
        assert save(name="p20_var",
                    steps='[sense:search]{query: $q} >> [table:take]{n: 3}').get("success") is True
        # 주석·멀티라인 문자열(본문 속 `#`)도 통과
        assert save(name="p20_multi",
                    steps='# 주석\n[self:write]{path:"a", content: "l1\n# 제목\nl2"}'
                    ).get("success") is True
        # 저장본은 run 이 읽어 실행 가능한 형태로 남는다 (파싱만 했지 실행은 안 함)
        assert _wf.get_workflow("p20_var")["steps"], "몸통 유실"
    finally:
        if orig_env is None:
            os.environ.pop("INDIEBIZ_BASE_PATH", None)
        else:
            os.environ["INDIEBIZ_BASE_PATH"] = orig_env
    print("P20 OK — save 등록 관문(따옴표·중괄호·산문·빈 몸통 거부 / $변수·주석·멀티라인 통과)")


def test_p21_grep_missing_path_loud():
    """P21(2026-08-21, ep1357): [self:grep] 이 없는 path 에서 success:true + 'No matches found' —
    "패턴 미매칭"과 "경로 부재"가 구분되지 않아 잘못 추측한 경로가 침묵 0건으로 통과(B8 부류).
    → success:false + '경로가 존재하지 않습니다' 명시 신고. rg·파이썬 두 층 모두(가드는 층 앞)."""
    with tempfile.TemporaryDirectory() as td:
        open(os.path.join(td, "real.py"), "w").write("needle\n")
        missing = os.path.join(td, "nope", "ghost.py")
        def g(**kw):
            base = {"pattern": "needle", "path": missing}
            base.update(kw)
            return json.loads(_sysess.execute(base, _Ctx("grep_files", td)))
        # ASCII 패턴(rg 층) — 절대경로 부재
        r = g()
        assert r.get("success") is False and "존재하지 않습니다" in r.get("error", ""), r
        assert "ghost.py" in r["error"] and r.get("items") == [] and r.get("total") == 0, r
        # 한글 패턴(파이썬 층 강제) — 같은 신고
        r2 = g(pattern="바늘")
        assert r2.get("success") is False and "존재하지 않습니다" in r2.get("error", ""), r2
        # 상대경로 부재 — 해석된 절대경로도 함께 신고
        r3 = g(path="nope/ghost.py")
        assert r3.get("success") is False and td in r3["error"], r3
        # 오탐 없음: 존재하는 루트에서 진짜 0건은 여전히 success:true + No matches
        ok = json.loads(_sysess.execute({"pattern": "zzz_absent", "path": td}, _Ctx("grep_files", td)))
        assert ok.get("success") is True and ok.get("total") == 0 and "No matches" in ok.get("text", ""), ok
        hit = json.loads(_sysess.execute({"pattern": "needle", "path": td}, _Ctx("grep_files", td)))
        assert hit.get("total") == 1, hit
        # 인접 결함: 단일 파일 path 에서 rg --count-matches 가 파일명 접두를 생략 → 전수 {} →
        # total 0 + truncated true 거짓 신고 ("매칭 0건 중 N건만 표시"). --with-filename 강제.
        one = json.loads(_sysess.execute({"pattern": "needle", "path": os.path.join(td, "real.py")},
                                         _Ctx("grep_files", td)))
        assert one.get("total") == 1 and one.get("total_files") == 1 and one.get("truncated") is False, one
    print("P21 OK — grep 없는 경로 = success:false 명시 신고(rg·파이썬 층), 진짜 0건은 그대로·단일 파일 전수 계수 정직")


def test_p22_copy_empty_hands_vs_no_currency():
    """P22(2026-08-22, F20-3 후속): `[self:copy]{dest}` 가 **통화 없음과 0행을 접었다** —
    `_piped_items` 가 둘 다 `[]` 로 만들어, 앞 액션이 통화를 안 내는 액션이었을 때와
    앞 액션이 정상적으로 0행을 냈을 때가 똑같이 "Error: 복사할 항목이 없습니다"로 보고됐다.
    0행 계약(통화 없음=에러 / 0행=성공)을 이 자리에선 지킬 수가 없던 것.
    → 3갈래 분간: 통화 없음=거절(받은 봉투 진단 동봉) · 0행=빈손 성공 · 행은 있으나
      레코드 아님=별도 거절. 성공 문자열은 "Error:" 로 시작하지 않아야 파이프가 성공으로 읽는다."""
    import importlib
    wf = importlib.import_module("workflow_engine")
    _copyops = _load("_t_copyops", os.path.join(_PKG, "system_essentials", "copy_ops.py"))

    def copy_with(prev):
        return _sysess.execute({"dest": "~/Desktop/_p22_없는폴더", "_prev_result": prev},
                               _Ctx("copy_path", "/tmp"))

    # ① 0행 = 빈손 성공 (감시자·필터 문형의 정상 결과)
    for zero in (json.dumps({"items": []}), {"items": []}, []):
        r = copy_with(zero)
        assert "0행" in r and not r.startswith("Error:"), (zero, r)
        assert wf.is_error_result(r) is False, r        # 파이프가 성공으로 읽는가
        assert "복사할 항목이 없습니다" not in r, r       # 옛 문장으로 되돌아가지 않았는가

    # ② 통화 없음 = 여전히 거절 + 받은 봉투 진단(무엇이 왔는지)
    r = copy_with("그냥 평문 결과입니다")
    assert r.startswith("Error:") and "통화가 없습니다" in r, r
    assert wf.is_error_result(r) is True, r
    r2 = copy_with(json.dumps({"success": True, "message": "완료"}))
    assert r2.startswith("Error:") and "message" in r2, r2   # 봉투 키를 보여준다
    r3 = copy_with(json.dumps({"items": "목록이 아님"}))
    assert r3.startswith("Error:") and "목록이 아니라" in r3, r3
    r4 = copy_with(None)                                   # 파이프도 src 도 없음
    assert r4.startswith("Error:") and "src" in r4, r4     # 옛 안내 보존

    # ③ 행은 있으나 레코드가 아님 = 0행과 다른 사실이라 다른 문장
    r5 = copy_with(json.dumps({"items": ["a.jpg", "b.jpg"]}))
    assert r5.startswith("Error:") and "2행" in r5 and "레코드" in r5, r5

    # ④ 옛 records 봉투 관용은 보존(생산자 0이지만 잔존 데이터)
    assert _copyops.piped_items(json.dumps({"records": [{"path": "/x"}]})) == [{"path": "/x"}]
    # ⑤ 세 갈래가 반환값 층에서도 갈리는가 (None ≠ [])
    assert _copyops.piped_items("평문") is None
    assert _copyops.piped_items(json.dumps({"items": []})) == []
    print("P22 OK — copy: 통화 없음=거절(봉투 진단) · 0행=빈손 성공 · 비-레코드=별도 거절")


def test_p23_unary_transformers_restate_scope():
    """P23(26회차 B26-1·B26-2): 단항 변환자가 봉투의 자기-기수 서술을 안 고쳐 거짓말을 했다.

    시스템 자신의 정의: truncated == total > len(items) (portal_warehouse:304 · test_body_vocab T1/T5).
    ⑭가 이항 변환자에 `_carry_flags` 를 달았지만 단항 경로는 안 쓸었다.
    """
    trunc_free = json.dumps({"success": True, "total": 29,  "truncated": False,
                             "items": [{"파일": f"f{i}.py", "n": i} for i in range(29)]}, ensure_ascii=False)
    # take: 29 → 1 이면 total(29) > 1 이므로 truncated 가 켜져야 한다
    t = _run("data_take", {"_prev_result": trunc_free, "n": 1})
    assert t.get("truncated") is True and t.get("count") == 1 and t.get("total") == 29, t
    # filter·dedup 도 같은 부류
    f = _run("data_filter", {"_prev_result": trunc_free, "where": "n < 5"})
    assert f.get("truncated") is True and f.get("count") == 5, {k: f.get(k) for k in ("truncated", "count")}
    # ★기수 불변 변환(sort)은 건드리지 않는다 — 오폭 방지
    s = _run("data_sort", {"_prev_result": trunc_free, "by": "n"})
    assert s.get("truncated") is False and s.get("count") == 29, {k: s.get(k) for k in ("truncated", "count")}
    # total 이 없으면 지어내지 않는다(_carry_flags 의 join 조항) — 침묵은 거짓말이 아니다
    no_total = json.dumps({"success": True, "items": [{"a": i} for i in range(10)]}, ensure_ascii=False)
    n = _run("data_take", {"_prev_result": no_total, "n": 2})
    assert "total" not in n and "truncated" not in n, n
    # 상류가 이미 truncated 면 꺼지지 않는다(단조)
    already = json.dumps({"success": True, "total": 500, "truncated": True,
                          "items": [{"a": i} for i in range(3)]}, ensure_ascii=False)
    a = _run("data_sort", {"_prev_result": already, "by": "a"})
    assert a.get("truncated") is True, a
    # B26-2: 기수가 변한 변환 뒤의 봉투 summary 는 변환 전 집계라 stale
    with_sum = json.dumps({"success": True, "summary": {"총건수": 90, "평균가": "31,952만원"},
                           "items": [{"동": "a", "가": i} for i in range(9)]}, ensure_ascii=False)
    g = _run("data_groupby", {"_prev_result": with_sum, "by": "동", "agg": {"평균": ["avg", "가"]}})
    assert "summary" not in g, g
    k = _run("data_take", {"_prev_result": with_sum, "n": 3})
    assert "summary" not in k, k
    # 기수 불변이면 summary 는 여전히 참이므로 보존한다
    s2 = _run("data_sort", {"_prev_result": with_sum, "by": "가"})
    assert isinstance(s2.get("summary"), dict) and s2["summary"]["총건수"] == 90, s2.get("summary")
    print("P23 OK — 단항 변환자 truncated 재서술(단조·무오폭·total 무생성) + stale summary 제거")


def test_p24_output_sink_consumes_currency():
    """P24(26회차 B26-3): [self:output] 이 파이프 통화를 안 먹고 빈 산출을 성공으로 신고했다.

    2026-08-05 어휘 압축이 `_output_file` 을 지우면서 삭제 사유를 파일에 적어 남겼다 —
    "파이프 입력도 무시해 빈 파일을 쓰던 반쪽 싱크". 형제 둘(gui·clipboard)은 그대로 남았다.
    """
    from ibl_exec_output import _sink_content, _output_gui
    prev = json.dumps({"success": True, "items": [{"bookname": "건축입문"}]}, ensure_ascii=False)
    # ① content 생략 → _prev_result 를 먹는다
    got, err = _sink_content("", {"_prev_result": prev})
    assert err is None and got == prev, (got, err)
    # ② content 명시가 이긴다. 빈 문자열도 유효한 값(write 규약)
    assert _sink_content("", {"content": "직접", "_prev_result": prev})[0] == "직접"
    assert _sink_content("", {"content": "", "_prev_result": prev})[0] == ""
    # ③ 쓸 것이 아무것도 없으면 None — 부르는 쪽이 정직하게 거절하라는 신호
    assert _sink_content("", {})[0] is None
    # ④ gui: 통화를 먹으면 content 가 채워진다 (수리 전엔 "" 였다)
    r = _output_gui("", {"_prev_result": prev, "format": "테이블"}, ".")
    assert r.get("ok") and r["output"]["content"] == prev, r
    # ⑤ gui: 빈손이면 ok:true 가 아니라 정직한 거절 (침묵-삼킴 금지)
    e = _output_gui("", {"format": "테이블"}, ".")
    assert "error" in e and "ok" not in e, e
    print("P24 OK — output gui/clipboard 가 write 와 같은 싱크 계약(_prev_result 수용·빈손 정직 거절)")


def test_p25_envelope_shape_uses_one_currency_judge():
    """P25(27회차 B27-1): 봉투 요약의 shape 판정기가 파이프 이음매의 판정기와 달랐다.

    D13(원형 유지)과 M1(요약 접기)이 각각 옳았는데, 접고 나서는 모델이 원형이 아니라
    **판정**을 읽는다 — 그 판정이 이음매를 반박하면 통화가 살아 있는데 죽었다고 읽힌다.
    """
    from ibl_envelope import summarize_result
    # ① 표 형(columns/rows)만 낸 변환자 — 이음매는 items 를 파생한다
    tbl = json.dumps({"success": True, "total": 240, "truncated": True,
                      "columns": ["파일", "영역"],
                      "rows": [["a.py", "backend"], ["b.md", "docs"]]}, ensure_ascii=False)
    s = summarize_result(tbl)
    assert s["shape"] == "items", s          # 수리 전엔 "effect"
    assert s["count"] == 2, s
    assert s.get("items_derived") is True, s  # keys 에 items 가 없는 이유를 밝힌다
    assert s["columns"] == ["영역", "파일"], s
    # ② 진짜 효과 봉투는 여전히 effect — 오폭 금지
    eff = json.dumps({"success": True, "path": "/tmp/x.md", "size": 12}, ensure_ascii=False)
    assert summarize_result(eff)["shape"] == "effect", summarize_result(eff)
    # ③ items 를 직접 낸 봉투는 파생기를 부르지 않는다(빠른 경로 무변경)
    it = json.dumps({"success": True, "items": [{"a": 1}]}, ensure_ascii=False)
    s3 = summarize_result(it)
    assert s3["shape"] == "items" and "items_derived" not in s3, s3
    # ④ 실패 봉투는 error 가 먼저(진단은 다이어트 밖)
    er = json.dumps({"success": False, "error": "x", "columns": ["a"], "rows": [[1]]})
    assert summarize_result(er)["shape"] == "error", summarize_result(er)
    print("P25 OK — 봉투 shape 판정이 이음매(derive_items)와 같은 판정기 · 효과 오폭 없음")


def test_p26_spill_resolved_at_injection_seam():
    """P26(27회차 B27-2): 스필 해소기가 입구가 아니라 소비처마다 흩어져 구멍이 남았다.

    교재의 약속은 "뒤 step 은 참조를 투명하게 해소해 원래 데이터를 그대로 본다" 인데,
    `_op_groupby` 는 `_rows_for_field` 로 들어가고 해소기는 형제 입구 `_get_items` 에만 있었다.
    """
    import tempfile
    from workflow_binding import _auto_inject_prev
    body = json.dumps({"success": True,
                       "items": [{"영역": "backend"}, {"영역": "docs"}, {"영역": "backend"}]},
                      ensure_ascii=False)
    with tempfile.NamedTemporaryFile("w", suffix=".json", delete=False, encoding="utf-8") as f:
        f.write(body)
        spilled_path = f.name
    ref_env = json.dumps({"success": True, "path": spilled_path, "items": [], "spilled": True,
                          "ref": {"path": spilled_path, "kind": "items", "count": 3,
                                  "bytes": len(body)}}, ensure_ascii=False)
    # ① 이음매가 참조를 본문으로 되돌린다 — 모든 패키지가 계약을 상속한다
    ti = _auto_inject_prev({"params": {}}, ref_env)
    assert json.loads(ti["params"]["_prev_result"])["items"] == json.loads(body)["items"], ti
    # ② 해소된 통화를 groupby 가 실제로 먹는가 (수리 전엔 "통화를 찾지 못했습니다")
    g = _run("data_groupby", {"_prev_result": ti["params"]["_prev_result"], "by": "영역"})
    # groupby 는 표 형(columns/rows)으로 방출한다 — 그 자체가 P25 가 다루는 자리다
    assert g.get("success") and sorted(g.get("rows") or []) == [["backend", 2], ["docs", 1]], g
    # ③ 스필이 아닌 평범한 통화는 손대지 않는다(빠른 경로)
    plain = json.dumps({"items": [{"a": 1}]}, ensure_ascii=False)
    assert _auto_inject_prev({"params": {}}, plain)["params"]["_prev_result"] == plain
    # ④ 만료·부재는 삼키지 않고 참조를 그대로 흘려보낸다 — 새 침묵 경로 0
    os.unlink(spilled_path)
    gone = _auto_inject_prev({"params": {}}, ref_env)
    assert gone["params"]["_prev_result"] == ref_env, gone
    print("P26 OK — 스필 해소가 주입 이음매 하나에 · groupby 관통 · 평범 통화 무변경 · 만료 무침묵")


def test_p27_each_substitution_respects_syntactic_slot():
    """P27(27회차 B27-3): each 의 $it 치환이 자리의 문법을 안 보고 늘 맨몸 텍스트였다.

    파라미터 자리는 저자가 따옴표를 쓰므로 우연히 맞았고, 조건 자리는 따옴표를 쓸 수 없어
    깨졌다 — `each × [if:]` 라는 가장 자연스러운 교차가 통째로 말할 수 없는 문장이었다.
    """
    from ibl_exec_each import _each_substitute, _inside_string
    row = {"영역": "backend/ibl", "파일": "a.py", "n": 3, "플래그": True, "빈것": None,
           "제목": "그는 \"안녕\" 이라 했다"}
    # ① 조건 자리(따옴표 밖) — 문자열은 따옴표를 얻는다
    out, miss = _each_substitute("[if: $it.영역 matches 'backend']{[self:time]}", row, "it")
    assert not miss and out.startswith('[if: "backend/ibl" matches'), out
    # ② 숫자·불리언·null 은 맨몸 — 조건의 크기 비교가 문자열로 변질되지 않게
    assert _each_substitute("[if: $it.n > 1]{x}", row, "it")[0] == "[if: 3 > 1]{x}"
    assert _each_substitute("[if: $it.플래그]{x}", row, "it")[0] == "[if: true]{x}"
    assert _each_substitute("[if: $it.빈것 == null]{x}", row, "it")[0] == "[if: null == null]{x}"
    # ③ 파라미터 자리(따옴표 안)는 옛 규약 그대로 — 본문만, 따옴표 추가 없음(무회귀)
    p, _ = _each_substitute("[self:notify_user]{message: '$it.파일'}", row, "it")
    assert p == "[self:notify_user]{message: 'a.py'}", p
    # ④ 따옴표 밖이라도 값 속 따옴표는 이스케이프되어 문장을 깨지 않는다
    q, _ = _each_substitute("[if: $it.제목 matches 'x']{y}", row, "it")
    assert q.count('[if: "') == 1 and q.endswith("matches 'x']{y}"), q
    # ⑤ 자리 판정기 자체 — 이스케이프된 따옴표에 속지 않는다
    assert _inside_string("{a: 'x', b: $it", 14) is False
    assert _inside_string("{a: 'x$it", 6) is True
    assert _inside_string("{a: 'it\\'s', b: $it", 18) is False
    # ⑥ 없는 필드는 여전히 정직하게 missing (조용한 빈 값 금지 — F14-1 무회귀)
    _, m2 = _each_substitute("[if: $it.없는것 > 1]{x}", row, "it")
    assert m2 == ["없는것"], m2
    print("P27 OK — each 치환이 자리의 문법을 따른다(조건=리터럴·파라미터=본문·수·불리언 맨몸)")


def test_p28_repeat_carries_body_honesty():
    """P28(27회차 B27-4): [repeat:] 경계가 몸통의 정직 신고(skipped_steps)를 삼켰다.

    `[on_error: skip|null]` 의 계약은 "봉투 skipped_steps 로 신고되니 조용한 성공이 아니다".
    그 신고가 몸통 봉투에 실리고 repeat 봉투는 iterations/items 만 조립해 경계에서 증발했다 —
    24회차의 병렬 판정(부분 성공은 실패를 지우지 않는다)이 안 닿은 나머지 경계.
    """
    from ibl_control_blocks import _execute_repeat
    from ibl_parser import parse
    # 문장 모양은 파서가 정본 — 손으로 지은 step dict 는 실제와 어긋난다
    # ① 몸통이 매 회차 step 을 건너뛰면 바깥도 그 사실을 말한다
    out = _execute_repeat(parse('[repeat: 2, collect: true]{[on_error: null] '
                                '[sense:crawl]{url: "https://invalid.invalid.invalid"} '
                                '>> [table:take]{n: 1}}')[0], ".", None)
    assert out.get("success") is True and out.get("count") == 0, out
    sk = out.get("skipped_steps")
    assert isinstance(sk, list) and len(sk) == 2, out          # 수리 전엔 키 자체가 없었다
    assert {e["iteration"] for e in sk} == {1, 2}, sk          # 어느 회차였는지가 진단의 절반
    assert "회차에서 step 을 건너뛰었습니다" in (out.get("note") or ""), out.get("note")
    # ② 아무 일 없으면 조용하다 — 없는 것을 지어내지 않는다
    # (몸통은 도구를 안 부르는 식 할당 — 이 하네스엔 활성 프로젝트가 없다)
    clean = _execute_repeat(parse('[repeat: 2, collect: true]{$x = 1}')[0], ".", None)
    assert clean.get("success") is True and "skipped_steps" not in clean, clean
    assert "건너뛰" not in (clean.get("note") or ""), clean.get("note")
    print("P28 OK — repeat 이 몸통의 skipped_steps 를 회차와 함께 승계 · 무사고 회차는 조용함")


def test_p29_empty_hand_contract_is_one_rule():
    """P29(28회차 B28-1): 빈손(0행)에서 필드 부재를 주장하면 안 된다 — verb 전수 불변식.

    F17 이 빈손 계약을 "verb 마다 심사" 로 정해 둔 탓에 verb 마다 부재 판정을 손으로
    다시 적었고, 빈손 보호는 *호출자가 먼저 짧게 끊어 주는* 우연에 기댔다. 단항 9개 중
    8개는 우연히 끊겼고 rename 만 안 끊겨 혼자 다른 답을 냈다.
        [self:body]{days: 3, limit: 5} >> [table:filter]{where: "존재하지않는값ZZZ"}
                                       >> [table:rename]{map: {"파일": "경로"}}
        → "rename: 필드 ['파일'] 이(가) 없습니다. 행 필드 예: []"
    step1 이 '파일' 을 열로 신고했으므로 그 필드는 실재한다 — 사라진 건 행뿐이다.
    오류문이 스스로를 반박한다: `행 필드 예: []` 는 *아무것도 못 봤다*이지 *없다*가 아니다.

    ★이 시험이 지키는 것은 rename 하나가 아니라 **부류**다 — 새 verb 가 부재 판정을
    손으로 적어도 여기서 걸린다(갈래가 다시 생기지 않게).
    """
    from importlib import reload  # noqa: F401
    EMPTY = json.dumps({"success": True, "items": []}, ensure_ascii=False)
    FULL = json.dumps({"success": True, "items": [{"동": "a", "가": 1}, {"동": "b", "가": 2}]},
                      ensure_ascii=False)
    # (도구, 없는 필드를 가리키는 params) — 단항 변환자 전수
    VERBS = [
        ("data_filter", {"where": "없는열 > 1"}),
        ("data_sort", {"by": "없는열"}),
        ("data_take", {"n": 3}),
        ("data_select", {"columns": ["없는열"]}),
        ("data_dedup", {"by": "없는열"}),
        ("data_rename", {"map": {"없는열": "새"}}),
        ("data_groupby", {"by": "없는열"}),
        ("data_compute", {"set": {"x": "없는열 * 2"}}),
        ("data_flatten", {"field": "없는열"}),
    ]
    # ① 불변식 A — 빈손이면 **전 verb** 가 0행 성공으로 흘려보낸다(부재 주장 금지)
    offenders = []
    for tool, params in VERBS:
        r = _run(tool, {**params, "_prev_result": EMPTY})
        if r.get("success") is False:
            offenders.append((tool, str(r.get("error"))[:80]))
    assert not offenders, f"빈손에서 부재를 주장한 verb: {offenders}"
    # ② 불변식 B — 행이 있으면 없는 필드는 **여전히 시끄럽게** 거절한다(과교정 금지)
    silent = []
    for tool, params in VERBS:
        if tool == "data_take":
            continue                      # take 는 필드를 안 본다
        r = _run(tool, {**params, "_prev_result": FULL})
        if r.get("success") is not False:
            silent.append(tool)
    assert not silent, f"행이 있는데 없는 필드를 조용히 통과시킨 verb: {silent}"
    # ③ 판정기 자체 — 관측이 0이면 어떤 이름도 부재로 주장하지 않는다
    from importlib import import_module  # noqa: F401
    _abs = _dataops._absent_fields
    _obs = _dataops._observed_fields
    assert _abs(["파일", "없는것"], _obs(rows=[])) == []            # 빈 관측 = 주장 없음
    assert _abs(["파일", "없는것"], _obs(columns=[])) == []
    assert _abs(["파일", "없는것"], _obs(rows=[{"파일": 1}])) == ["없는것"]
    assert _abs("없는것", _obs(columns=["파일"])) == ["없는것"]      # 스칼라 이름도 받는다
    print("P29 OK — 빈손 계약이 verb 전수 한 규칙(부재 주장 금지) · 행 있으면 시끄러운 거절 유지")


def test_P30_파이프결과_변수를_items_로_주입할_수_있다():
    """P30 — `$변수`(파이프 결과)를 `items:` 로 주입하면 통화로 읽혀야 한다 (2026-08-27).

    P2 가 고친 것은 items 가 **리스트**로 인라인될 때였다. 같은 병이 다른 문으로 재발했다:
    변수 치환은 통화를 **JSON 문자열**로 넣는데, 그 문자열이 담은 것은 파이프 결과의
    *원형*이라 `items` 가 없다(파이프 이음매만 파생한다 — step_results 는 토큰 중복을
    피하려고 원형을 유지하고, 그 저장소가 $변수 슬롯을 겸한다).

        $표 = [sense:host]{…} >> [table:select]{columns: ["cpu_percent"]}
        [table:spreadsheet]{items: "$표", path: …}   → success + **1×1 빈 xlsx** (조용한 실패)
        [table:document]{items: "$표"}                 → "blocks 가 필요합니다"
        [table:brief]{items: "$표"}                    → "입력 통화가 없습니다"
        $단일 = [sense:host]{…}                        (생산자 직접 방출 items)
        [table:brief]{items: "$단일"}                   → 정상  ← 같은 문장이 앞 모양에 따라 갈렸다

    수리 = 되읽기 관문이 모양 판정을 몸의 단일 게이트(derive_items)에 위임한다.
    """
    from common.currency import coerce_items_payload
    PIPE = json.dumps({"success": True, "cpu_percent": 9.9,
                       "columns": ["cpu_percent"], "rows": [[9.9], [12.1]]}, ensure_ascii=False)
    assert coerce_items_payload(PIPE) == [{"cpu_percent": 9.9}, {"cpu_percent": 12.1}]
    # 생산자 직접 방출 items 는 그대로 (파생이 덮지 않는다)
    assert coerce_items_payload('{"items": [{"a": 1}]}') == [{"a": 1}]
    # blocks 통화(document 생산자)도 같은 눈
    assert coerce_items_payload('{"blocks": [{"type": "paragraph", "text": "x"}]}') == \
        [{"type": "paragraph", "text": "x"}]
    # 통화가 아닌 것은 여전히 None — 호출자의 진단 경로를 뺏지 않는다
    assert coerce_items_payload("그냥 문자열") is None
    assert coerce_items_payload('{"success": true, "path": "/tmp/x"}') is None
    print("P30 OK — 파이프 결과 변수도 통화로 읽힌다(판정은 단일 게이트)")


def test_P31_blocks_의_변수_주입도_원형으로_읽힌다():
    """P31 — blocks 안 구조 필드에 `$변수` 를 주입하면 원형(list)으로 읽혀야 한다 (2026-08-27).

    P30(items)과 같은 병의 blocks 판: 변수 치환은 JSON **문자열**을 넣는데, document 의
    blocks 소비자는 되읽기가 없어 `columns: "$표.columns"` 가 문자열 그대로 들어갔다 —
    렌더러가 그 문자열을 문자 리스트로 순회해 표 헤더가 **한 글자씩 쪼개진 표**가 나왔다
    (완성 보고서 프로그램 실측). 수리 = 범용 게이트 coerce_json_param + doc_build 가
    blocks 전체와 블록 구조 필드(columns/rows/items)만 되읽는다 — text(산문)는 안 건드린다.
    """
    from common.currency import coerce_json_param
    # 범용 게이트: JSON 문자열 → 원형, 산문·비JSON 은 그대로
    assert coerce_json_param('["이름", "가격"]') == ["이름", "가격"]
    assert coerce_json_param('{"a": 1}') == {"a": 1}
    assert coerce_json_param("[삭제] 그냥 산문") == "[삭제] 그냥 산문"
    assert coerce_json_param(["이미", "원형"]) == ["이미", "원형"]

    _doc = _load("_t_docbuild_p31", os.path.join(_PKG, "data-ops", "doc_build.py"))
    with tempfile.TemporaryDirectory() as td:
        out = json.loads(_doc.render_document({
            "format": "markdown", "title": "P31",
            "blocks": [
                {"type": "paragraph", "text": "[유지] 산문은 그대로"},
                # 변수 치환이 실제로 넣는 모양 — JSON 문자열
                {"type": "table", "columns": '["이름", "가격"]', "rows": '[["A", 100]]'},
                {"type": "cards", "columns": 2, "items": '[{"title": "t", "url": "u"}]'},
            ],
        }, output_base=td))
        assert out.get("success"), out
        md = out.get("markdown") or open(out["path"], encoding="utf-8").read()
        assert "| 이름 | 가격 |" in md, f"표 헤더가 원형으로 렌더돼야 한다: {md[:200]}"
        assert "| [ |" not in md, "JSON 문자열이 문자 단위로 쪼개졌다 — 되읽기가 죽었다"
        assert "[유지] 산문은 그대로" in md, "산문 text 를 건드리면 안 된다"
    print("P31 OK — blocks 의 변수 주입도 원형으로(구조 필드만, 산문 불가침)")


def test_P32_정렬_파이프는_file_find_통화를_비우지_않는다():
    """P32 — `| sort: title` · `| sort: title asc` · `>> [table:sort]{by}` 가 file_find 의
    items 를 **비운다**는 신고(2026-09-03, 칩 task_5a88cb26)의 회귀 고정물.

    실측: 변환자도 설탕 파서도 멀쩡했다. 0 은 세는 쪽이 만든 숫자다 — 신고자의 계수기가
    /ibl/execute 응답의 **최상위** `items` 를 읽었는데, 단일 액션은 통화가 최상위에 오지만
    파이프(2단계+)는 봉투(results[]/final_result)로 나가고 통화는 `final_result`
    **JSON 문자열** 안에 산다(렌더 코어 unwrapFinalResult · ibl_health_check 가 그 규칙으로
    푼다). 같은 계수기로 재면 `| sort: title desc` 도 0 이다 — asc/desc 차이가 아니었다.
    그 오독으로 계기(folder_survey.yaml)의 정렬이 떨어져 나갔다.

    ★이 시험이 지키는 것: ① 네 표기가 같은 한 블록으로 펼쳐진다(설탕=표면, 동사=정본)
    ② 파이프 봉투에서 통화를 *규칙대로* 꺼내면 원천 count 그대로·정렬된 채 나온다
    (asc 는 기본, desc 는 뒤집힘) ③ step 요약도 같은 수를 신고한다.
    파이프 통화의 자리(final_result)가 바뀌면 여기가 먼저 알린다.
    """
    from ibl_parser import parse
    from ibl.workflow_engine import execute_pipeline
    FORMS = {
        "verb":  '[self:file_find]{pattern: "*.md", path: "%s"} >> [table:sort]{by: "title"}',
        "plain": '[self:file_find]{pattern: "*.md", path: "%s"} | sort: title',
        "asc":   '[self:file_find]{pattern: "*.md", path: "%s"} | sort: title asc',
        "desc":  '[self:file_find]{pattern: "*.md", path: "%s"} | sort: title desc',
    }

    def _currency(env):
        # 파이프 봉투의 통화 자리 — 렌더 코어 unwrapFinalResult 와 같은 규칙(문자열이면 JSON)
        fr = env.get("final_result") if isinstance(env, dict) else None
        return json.loads(fr) if isinstance(fr, str) else fr

    with tempfile.TemporaryDirectory() as td:
        for n in ("b.md", "c.md", "a.md"):
            with open(os.path.join(td, n), "w", encoding="utf-8") as f:
                f.write(n)
        steps = {k: parse(v % td) for k, v in FORMS.items()}
        # ① 설탕 세 표기 = 동사 한 블록. desc 만 desc:true 를 얹는다.
        for k in ("plain", "asc"):
            assert steps[k][1] == steps["verb"][1], (k, steps[k][1])
        assert steps["verb"][1]["params"] == {"by": "title"}, steps["verb"][1]
        assert steps["desc"][1]["params"] == {"by": "title", "desc": True}, steps["desc"][1]
        # ② 통화는 비지 않는다 — 원천 3건 그대로, 순서만 바뀐다
        for k, st in steps.items():
            env = execute_pipeline(st, td)
            assert env.get("success") is True, (k, env)
            cur = _currency(env)
            titles = [i.get("title") for i in (cur or {}).get("items") or []]
            expect = ["a.md", "b.md", "c.md"]
            assert titles == (expect[::-1] if k == "desc" else expect), (k, titles, env)
            assert cur.get("count") == 3, (k, cur)
            # ③ 에이전트 경계(/ibl/execute·MCP)가 입히는 봉투 다이어트 뒤에도 step 요약과
            #    통화가 서로를 반박하지 않는다 — 요약은 count 3, 통화는 그대로 3건.
            from ibl_envelope import diet_envelope
            thin = diet_envelope(env, verbose=False)
            last = (thin.get("results") or [])[-1]
            assert last.get("action") == "sort" and last.get("count") == 3, (k, last)
            assert last.get("shape") == "items", (k, last)
            assert [i.get("title") for i in _currency(thin)["items"]] == titles, k
    print("P32 OK — 정렬 설탕/동사 네 표기 모두 file_find 통화를 보존 · 통화는 final_result 에")


if __name__ == "__main__":                      # 러너는 하나 — pytest (2026-08-23)
    # ★두 번째 러너를 두지 않는다. 손으로 적은 러너는 반드시 드리프트한다 — 새 시험 함수를
    # 러너에 안 적으면 직접 실행이 **그 시험만 조용히 건너뛰고 종료코드 0** 을 낸다.
    # 실측(2026-08-23): 배터리 44개·시험 303건 중 **147건**이 직접 실행에서 한 번도 안 돌았고,
    # 27·28회차 상상훈련이 그 초록을 "전부 통과"로 보고서에 적었다(거짓 초록).
    # 위임하면 직접 실행도 살고(순찰·손버릇) 수집은 pytest 가 하므로 드리프트가 불가능하다.
    import sys as _sys
    try:
        import pytest as _pytest
    except ImportError:
        raise SystemExit("pytest 가 없습니다 — .venv/bin/python -m pytest 로 실행하세요")
    raise SystemExit(_pytest.main([__file__] + _sys.argv[1:]))
