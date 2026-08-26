"""[engines:render]{op:"xlsx"} 계약 배터리 — RENDER_XLSX_HANDOFF (2026-08-27).

★심장 = 재계산 관문: LibreOffice 는 xlsx 수식을 기본 재계산하지 않는다(캐시값 사용).
openpyxl 로 만든 fixture 는 수식 캐시가 없으므로, 프로파일 시딩(OOXMLRecalcMode=0)이
안 먹으면 PDF 에 계산값이 아예 없다 — "낡은 숫자가 찍힌 그림을 관찰했다고 믿는 침묵"을
이 관문이 기계적으로 실패시킨다 (2026-08-26 실측: 시딩 시 PDF 텍스트 '2\\n3\\n5' GREEN).

soffice 없으면 재계산 관문만 skip (test_render_core 의 node 부재 관례) — 정직 실패·
라우팅 계약은 soffice 없이도 돈다.

★이 배터리는 backend/ 밖(data/packages)의 파일을 읽는다 — 라이브 트리에서 돌릴 것.
"""
import json
import os
import sys

import pytest

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
import boot_paths  # noqa: F401 — 층 디렉토리 등재

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
RA = os.path.join(ROOT, "data/packages/installed/tools/media_producer/render_artifact.py")
HANDLER = os.path.join(ROOT, "data/packages/installed/tools/media_producer/handler.py")


def _load(path, name):
    import importlib.util
    spec = importlib.util.spec_from_file_location(name, path)
    mod = importlib.util.module_from_spec(spec)
    spec.loader.exec_module(mod)
    return mod


@pytest.fixture(scope="module")
def ra():
    return _load(RA, "mp_render_artifact_test")


@pytest.fixture(scope="module")
def fixture_xlsx(tmp_path_factory):
    """캐시 없는 수식 장부 — openpyxl 은 수식 캐시값을 쓰지 않는다(관문의 재료)."""
    import openpyxl
    p = tmp_path_factory.mktemp("ledger") / "fixture.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws["A1"] = 2
    ws["A2"] = 3
    ws["A3"] = "=A1+A2"
    wb.save(str(p))
    return str(p)


@pytest.fixture(scope="module")
def rendered(ra, fixture_xlsx, tmp_path_factory):
    """soffice 1회 실행을 모듈 전체가 공유 (변환 ~수 초 — 실행 수 최소화)."""
    if not ra._find_soffice():
        pytest.skip("LibreOffice(soffice) 없음 — 재계산 관문 생략")
    out_dir = str(tmp_path_factory.mktemp("render_out"))
    mtime_before = os.path.getmtime(fixture_xlsx)
    result = json.loads(ra.render_op_xlsx(
        {"path": fixture_xlsx, "viewports": ["1280x720"]}, output_base=out_dir))
    return {"result": result, "src": fixture_xlsx, "mtime_before": mtime_before}


def test_recalc_gate_pdf_has_computed_value(rendered):
    """★관문: 캐시 없는 =A1+A2 가 PDF 텍스트에 5 로 실재해야 한다 — 재계산의 유일한 증거."""
    result = rendered["result"]
    assert result.get("items"), f"변환 실패: {result}"
    pdf_path = result.get("pdf_path")
    assert pdf_path and os.path.exists(pdf_path), "pdf_path 동봉 계약 위반"
    import pymupdf
    doc = pymupdf.open(pdf_path)
    text = "".join(page.get_text() for page in doc)
    doc.close()
    assert "5" in text, (
        f"재계산 미작동 — PDF 텍스트에 계산값 5 부재: {text!r}. "
        "프로파일 OOXMLRecalcMode 시딩이 안 먹은 것(낡은 캐시 침묵 부류).")


def test_rows_are_xlsx_pngs(rendered):
    """행 계약: op=xlsx, path=실재하는 PNG 절대 경로 (GoalEval 시각 수집기 연동 조건)."""
    for row in rendered["result"]["items"]:
        assert row["op"] == "xlsx"
        assert os.path.isabs(row["path"]) and os.path.exists(row["path"])
        assert row["path"].endswith(".png")


def test_source_untouched(rendered):
    """지각 순수성: 원본 장부 불변 — 재계산은 LibreOffice 메모리 안에서만."""
    assert os.path.getmtime(rendered["src"]) == rendered["mtime_before"]


def test_viewports_ignored_with_note(rendered):
    """조용한 무시 금지: xlsx 에 viewports 를 주면 note 로 비적용을 알린다."""
    assert "viewports" in rendered["result"].get("note", "")


def test_korean_glyphs_not_silently_lost(ra, tmp_path_factory):
    """★한글 잉크 관문 (2026-08-27 실측): 맥 헤드리스 LibreOffice 는 시스템 폰트 폴백이 죽어
    한글이 텍스트 층에만 있고 픽셀에서 조용히 증발했다. 수리 = 프로파일 user/fonts 에
    CJK 시스템 폰트 링크. 이 관문은 한글만 든 셀이 실제 잉크로 그려졌는지를 픽셀로 잰다 —
    폰트명·OS 에 무관한 판정(글리프 유실이면 해당 영역이 전부 백지)."""
    if not ra._find_soffice():
        pytest.skip("LibreOffice(soffice) 없음 — 한글 잉크 관문 생략")
    import openpyxl
    d = tmp_path_factory.mktemp("hangul")
    src = d / "hangul.xlsx"
    wb = openpyxl.Workbook()
    wb.active["A1"] = "품목합계"                     # 한글만 — 숫자·라틴 없음
    wb.save(str(src))
    result = json.loads(ra.render_op_xlsx({"path": str(src)}, output_base=str(d)))
    assert result.get("items"), f"변환 실패: {result}"
    import pymupdf
    pix = pymupdf.Pixmap(result["items"][0]["path"])
    # 상단 1/4 영역에 잉크(비백색 픽셀)가 있어야 한다 — A1 밖은 전부 빈 시트다.
    stride = pix.width * pix.n
    top = pix.samples[: stride * (pix.height // 4)]
    assert min(top) < 200, "한글 글리프 유실 — 페이지 상단이 전부 백지(폰트 폴백 주입 실패)"


def test_soffice_absent_honest_error(ra, fixture_xlsx, tmp_path, monkeypatch):
    """soffice 부재 = 설치 명령을 담은 정직 실패 (B21-1 — 평문 오류 금지)."""
    monkeypatch.setattr(ra, "_find_soffice", lambda: None)
    result = json.loads(ra.render_op_xlsx({"path": fixture_xlsx}, output_base=str(tmp_path)))
    assert result["success"] is False
    assert "LibreOffice" in result["error"]
    assert ("brew install" in result["error"] or "libreoffice" in result["error"].lower())


def test_non_spreadsheet_rejected(ra, tmp_path):
    """.csv 등 타 확장자 = 정직 거절 + 대안 안내 (시각 형태가 내재하지 않은 데이터)."""
    p = tmp_path / "data.csv"
    p.write_text("a,b\n1,2\n")
    result = json.loads(ra.render_op_xlsx({"path": str(p)}, output_base=str(tmp_path)))
    assert result["success"] is False
    assert "xlsx/xlsm" in result["error"]


def test_missing_path_rejected(ra, tmp_path):
    result = json.loads(ra.render_op_xlsx({}, output_base=str(tmp_path)))
    assert result["success"] is False


class _Ctx:
    """execute() 스텁 컨텍스트 — 라우팅 계약 검증용."""
    tool_name = "render_artifact"

    def __init__(self, out):
        self._out = out

    def output_dir(self):
        return self._out


@pytest.fixture(scope="module")
def handler():
    return _load(HANDLER, "mp_handler_test")


def test_ext_inference_routes_xlsx(handler, fixture_xlsx, tmp_path):
    """Phase 1.5: op 생략 + .xlsx path → xlsx 갈래로 라우팅 (화면검수 문장 무수정 연동).

    soffice 를 없앤 채 호출 — LibreOffice 오류가 나오면 xlsx 갈래에 닿은 증거다
    (html 갈래였다면 바이너리를 그대로 렌더하려 들었을 것)."""
    fn = handler._OP_DISPATCHERS["render_artifact"]["xlsx"]
    orig = fn.__globals__["_find_soffice"]
    fn.__globals__["_find_soffice"] = lambda: None
    try:
        result = json.loads(handler.execute({"path": fixture_xlsx}, _Ctx(str(tmp_path))))
    finally:
        fn.__globals__["_find_soffice"] = orig
    assert result.get("success") is False and "LibreOffice" in result.get("error", "")


def test_ext_inference_default_html_without_path(handler, tmp_path):
    """path 없음 = 현행 기본 html 유지 (html 갈래의 입력 요구 오류가 그 증거)."""
    result = json.loads(handler.execute({}, _Ctx(str(tmp_path))))
    assert result.get("success") is False
    assert "html" in result.get("error", "")


if __name__ == "__main__":                      # 러너는 하나 — pytest (2026-08-23)
    try:
        import pytest as _pytest
    except ImportError:
        raise SystemExit("pytest 가 없습니다 — .venv/bin/python -m pytest 로 실행하세요")
    raise SystemExit(_pytest.main([__file__] + sys.argv[1:]))
