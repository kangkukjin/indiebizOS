"""[self:sheet] — 살아있는 엑셀 장부 부분 편집 (find / append / update).

시장 실측(ep951)의 수요 실체는 새 파일 생성이 아니라 **기존 장부**(재고·입출고·근태·연차)에
행을 더하고 셀을 고치는 일이다. [table:spreadsheet]는 통째 새로 쓰기라 기존 파일에 쓰면
수식·서식·다른 시트가 날아간다. 여기는 openpyxl 로 열어 부분만 만지고 원자 교체로 저장 —
수식·서식·병합·다른 시트가 보존된다. .xlsm 은 keep_vba 로 매크로도 보존.

★한계(가이드 sheet.md 명기): openpyxl 은 차트·이미지를 저장 시 유실한다. 차트가 든 장부는
데이터 시트와 차트 시트를 분리해 두는 것이 안전하다.

핸들러 계약: handler.py 의 op 디스패처가 tool_input 에 _project_path(상대경로 해석)와
_path_guard(쓰기 범위 검증 함수)를 주입해 부른다. 반환은 JSON 직렬화 가능한 dict.
"""
import os
import re as _re
from datetime import date, datetime, time
from pathlib import Path


# ── 공통 ──────────────────────────────────────────────────────────

def _resolve(tool_input):
    """(Path, err) — path 파라미터 해석. 상대경로는 _project_path 기준."""
    raw = tool_input.get("path") or tool_input.get("file_path") or ""
    if not raw:
        return None, {"success": False, "error": "path 가 필요합니다 — 대상 xlsx/xlsm 파일 경로."}
    p = Path(str(raw))
    if not p.is_absolute():
        p = Path(tool_input.get("_project_path") or ".") / p
    if not p.exists():
        return None, {"success": False, "error": f"파일을 찾을 수 없습니다: {p}"}
    if p.suffix.lower() not in (".xlsx", ".xlsm"):
        return None, {"success": False,
                      "error": f"xlsx/xlsm 만 지원합니다: {p.name} — 새 파일 생성은 [table:spreadsheet], 읽기는 [self:read]."}
    return p, None


def _open(p):
    import openpyxl
    # data_only=False(기본) — 수식 원문 보존이 이 액션의 존재 이유
    return openpyxl.load_workbook(str(p), keep_vba=(p.suffix.lower() == ".xlsm"))


def _pick_sheet(wb, tool_input):
    sn = tool_input.get("sheet")
    if sn:
        if sn not in wb.sheetnames:
            return None, {"success": False,
                          "error": f"시트가 없습니다: {sn} — 이 파일의 시트: {', '.join(wb.sheetnames)}"}
        return wb[sn], None
    return wb.active, None


def _headers(ws, header_row):
    """{헤더명: 열번호} — header_row 의 비어있지 않은 셀."""
    hmap = {}
    for c in range(1, (ws.max_column or 0) + 1):
        v = ws.cell(row=header_row, column=c).value
        if v is not None and str(v).strip():
            hmap[str(v).strip()] = c
    return hmap


def _last_data_row(ws, hmap, header_row):
    """매핑된 열 기준 마지막 값 행 — max_row 는 서식만 있는 유령 행도 세므로 직접 훑는다."""
    last = header_row
    for r in range(header_row + 1, (ws.max_row or header_row) + 1):
        if any(ws.cell(row=r, column=c).value is not None for c in hmap.values()):
            last = r
    return last


def _match(cell_val, want):
    """문자열 정규화 동치 + 숫자 동치(쉼표 허용) — '3500' == 3500 == '3,500'."""
    a = "" if cell_val is None else str(cell_val).strip()
    b = "" if want is None else str(want).strip()
    if a == b:
        return True
    try:
        return float(a.replace(",", "")) == float(b.replace(",", ""))
    except (ValueError, TypeError):
        return False


def _jsonable(v):
    # ★F53-3 (2026-09-02): 셀 타입 정규화는 read xlsx 와 **한 벌**(common.currency.normalize_cell)
    #   — 같은 셀을 두 독자가 다른 타입으로 내던 비대칭 봉합.
    try:
        from common.currency import normalize_cell
        return normalize_cell(v, none=None)
    except ImportError:
        if isinstance(v, (datetime, date, time)):
            return v.isoformat()
        return v


def _validate_keys(keys, hmap, what):
    unknown = [k for k in keys if k not in hmap]
    if unknown:
        return {"success": False,
                "error": f"{what}에 없는 열: {', '.join(unknown)} — 이 시트의 열: {', '.join(hmap)} "
                         f"(header_row 가 다르면 header_row 파라미터로 지정)"}
    return None


def _guard(tool_input, p):
    """쓰기 op 경로 가드 — handler 의 _validate_path_in_scope 주입분."""
    g = tool_input.get("_path_guard")
    if callable(g):
        err = g(str(p), tool_input.get("_project_path") or str(p.parent))
        if err:
            return {"success": False, "error": err}
    return None


def _save(wb, p):
    """원자 교체 저장 — 살아있는 장부를 반쪽 파일로 만들지 않는다."""
    tmp = p.with_name(p.name + ".tmp~")
    wb.save(str(tmp))
    os.replace(tmp, p)


def _prep(tool_input, need_where=False):
    """(ctx dict, err) — 공통 준비: 경로/워크북/시트/헤더/where 검증."""
    p, err = _resolve(tool_input)
    if err:
        return None, err
    try:
        header_row = int(tool_input.get("header_row", 1) or 1)
    except (TypeError, ValueError):
        header_row = 1
    wb = _open(p)
    ws, err = _pick_sheet(wb, tool_input)
    if err:
        wb.close()
        return None, err
    hmap = _headers(ws, header_row)
    if not hmap:
        wb.close()
        return None, {"success": False,
                      "error": f"{header_row}행에 헤더가 없습니다 — 헤더가 다른 행이면 header_row 로 지정."}
    where = tool_input.get("where") or {}
    if not isinstance(where, dict):
        wb.close()
        return None, {"success": False, "error": "where 는 {열이름: 값} 형태의 객체여야 합니다."}
    if need_where and not where:
        wb.close()
        return None, {"success": False,
                      "error": "update 는 where({열: 값})가 필수입니다 — 전행 갱신 사고 방지. 먼저 op:find 로 대상 행을 확인하세요."}
    err = _validate_keys(where.keys(), hmap, "where")
    if err:
        wb.close()
        return None, err
    return {"p": p, "wb": wb, "ws": ws, "hmap": hmap, "header_row": header_row, "where": where}, None


def _matched_rows(ctx):
    ws, hmap, where = ctx["ws"], ctx["hmap"], ctx["where"]
    last = _last_data_row(ws, hmap, ctx["header_row"])
    for r in range(ctx["header_row"] + 1, last + 1):
        if all(_match(ws.cell(row=r, column=hmap[k]).value, v) for k, v in where.items()):
            yield r


# ── op 구현 ───────────────────────────────────────────────────────

# 집계 배제 의심 신호(⑲, 2026-08-08 실험 10 → 실험 11 측정 라운드로 규칙 교정).
# 물음의 반전(보강④): "합계 행이 어디 있나"가 아니라 **"내가 넣을 행을 참조하지 않는
# 집계 수식이 있나"** — 바닥 합계(F1)·수동 덧셈(F3)·상단 합계(F4)·중간 소계(F5)·
# SUMPRODUCT(F6)를 한 물음으로 잡고, 가로 합계 열(F2 월계)은 자기-행 참조라 안 잡힌다.
# 판정 축=함수 이름이 아니라 **행**(실험 11 — 세로 집계 = 자기 행을 참조하지 않는 수식).
# ★배치 자동화(합계 위 삽입+참조 재작성)는 이 기준선(P19)에서 오탐0·미탐0 유지 후 별도 라운드.
_TOTAL_LABELS = {"합계", "총계", "소계", "총합", "계", "누계", "총액", "합계액",
                 "total", "totals", "grandtotal", "sum", "subtotal"}
# A1 참조 파서(단순) — 한계: 명명 범위·구조적 참조(Table[열])는 못 본다(실험 11 한계 명시).
# 행 절대참조($숫자) 캡처(실험 12 보강⑤) — 상수표 참조(=$B$2*$C$2)와 집계를 가른다.
_A1_REF = r"\$?[A-Z]{1,3}(\$?)(\d{1,7})(?:\s*:\s*\$?[A-Z]{1,3}(\$?)(\d{1,7}))?"
_QUALIFIED_RE = _re.compile(r"(?:'[^']+'|[A-Za-z0-9_가-힣.]+)!\s*" + _A1_REF)
_UNQUALIFIED_RE = _re.compile(r"(?<![A-Za-z0-9_$!])" + _A1_REF)

# 스캔 상한(⑳-1 — 실험 12): 상한 자체는 두되, 걸렸으면 반드시 신고한다.
_SCAN_MAX_SAME = 3000
_SCAN_MAX_OTHER = 1000


def _row_intervals(formula, sheet_name=None):
    """수식이 참조하는 행 구간 [(lo, hi, is_range, is_abs)].

    sheet_name=None → 무자격 참조만(같은 시트). sheet_name 지정 → 그 시트를
    자격 참조('시트'! 또는 시트!)하는 구간만 (교차 시트 — 실험 10 의 요약!COUNTA 부류).
    is_abs = 참조의 행 부분이 전부 절대($) — 보강⑤의 재료.
    """
    def _tup(m):
        lo, hi = int(m.group(2)), int(m.group(4) or m.group(2))
        is_range = m.group(4) is not None
        is_abs = bool(m.group(1)) and (not is_range or bool(m.group(3)))
        return (min(lo, hi), max(lo, hi), is_range, is_abs)

    if sheet_name:
        pat = _re.compile(r"(?:'" + _re.escape(sheet_name) + r"'|"
                          + _re.escape(sheet_name) + r")!\s*" + _A1_REF)
        return [_tup(m) for m in pat.finditer(formula)]
    s = _QUALIFIED_RE.sub(" ", formula)  # 자격 참조는 같은-시트 축에서 제외
    return [_tup(m) for m in _UNQUALIFIED_RE.finditer(s)]


def _is_aggregate(ivs, own_row=None):
    """세로 집계 판정(실험 11 후보 규칙 + 실험 12 보강⑤).

    자기 행을 참조하면 가로 합계(F2 월계) — 집계 아님.
    콜론 범위가 없고 전부 절대행이면 상수표 참조(=$B$2*$C$2 견적서형) — 집계 아님(⑤).
    ★⑤를 "전부 절대행이면 제외"로 넓히면 =SUM($D$3:$D$5) 절대범위 합계가 미탐이
    된다(실험 12 F10 실측) — 콜론 범위 부재 조건이 필수.
    나머지: 콜론 범위가 있거나 참조 행 수가 2 이상이면 집계.
    """
    if not ivs:
        return False
    if own_row is not None and any(lo <= own_row <= hi for lo, hi, _, _ in ivs):
        return False
    has_range = any(rng for _, _, rng, _ in ivs)
    if not has_range and all(abs_ for _, _, _, abs_ in ivs):
        return False  # ⑤
    span = sum(hi - lo + 1 for lo, hi, _, _ in ivs)
    return has_range or span >= 2


def _aggregates_missing_rows(wb, ws, header_row, new_lo, new_hi, cap=5):
    """새 행(new_lo~new_hi)을 참조하지 않는 세로 집계 수식들 — 전 시트 스캔(보강④).

    반환: (flags, scan_truncated) — 상한(⑳-1)에 걸려 못 본 구간이 있으면 True.
    상한 밖의 합계는 원리적으로 못 보므로, 그 사실을 침묵하지 않는 것이 계약이다
    (⑥′·⑭·⑰과 같은 부류 — 이 수리 코드 안에서 같은 병이 재발했던 것을 실험 12 가 잡았다).
    """
    flags = []
    scan_truncated = False

    def _covered(ivs):
        return any(lo <= new_hi and hi >= new_lo for lo, hi, _, _ in ivs)

    if ws.max_row > _SCAN_MAX_SAME:
        scan_truncated = True
    for row in ws.iter_rows(min_row=header_row + 1, max_row=min(ws.max_row, _SCAN_MAX_SAME)):
        for cell in row:
            v = cell.value
            if not (isinstance(v, str) and v.startswith("=")):
                continue
            ivs = _row_intervals(v)
            if _is_aggregate(ivs, own_row=cell.row) and not _covered(ivs):
                flags.append({"sheet": ws.title, "cell": cell.coordinate,
                              "row": cell.row, "formula": v[:60]})
                if len(flags) >= cap:
                    return flags, scan_truncated
    for other in wb.worksheets:
        if other is ws:
            continue
        if other.max_row > _SCAN_MAX_OTHER:
            scan_truncated = True
        for row in other.iter_rows(max_row=min(other.max_row, _SCAN_MAX_OTHER)):
            for cell in row:
                v = cell.value
                if not (isinstance(v, str) and v.startswith("=") and ws.title in v):
                    continue
                ivs = _row_intervals(v, sheet_name=ws.title)
                if _is_aggregate(ivs) and not _covered(ivs):
                    flags.append({"sheet": other.title, "cell": cell.coordinate,
                                  "row": cell.row, "formula": v[:60]})
                    if len(flags) >= cap:
                        return flags, scan_truncated
    return flags, scan_truncated


def _totals_label_row(ws, hmap, last_row, header_row):
    """보조 근거(라벨 축) — 마지막 행 첫 열이 합계류 라벨이면 그 행 번호.
    유니코드 공백(전각 U+3000 포함) 정규화 — 실험 11 F6 '합　계' 미탐 해소."""
    if last_row <= header_row:
        return None
    first = ws.cell(row=last_row, column=min(hmap.values())).value
    if isinstance(first, str) and "".join(first.split()).lower() in _TOTAL_LABELS:
        return last_row
    return None


def op_find(tool_input):
    """조건으로 행 검색 — where 생략 시 전체 (limit 기본 50). items 에 _row(행 번호) 포함."""
    ctx, err = _prep(tool_input)
    if err:
        return err
    try:
        limit = int(tool_input.get("limit", 50) or 50)
    except (TypeError, ValueError):
        limit = 50
    ws, hmap = ctx["ws"], ctx["hmap"]
    items, matched = [], 0
    for r in _matched_rows(ctx):
        matched += 1
        if len(items) < limit:
            item = {"_row": r}
            for name, c in hmap.items():
                item[name] = _jsonable(ws.cell(row=r, column=c).value)
            items.append(item)
    max_col = ws.max_column
    ctx["wb"].close()
    res = {"success": True, "sheet": ws.title, "columns": list(hmap),
           "matched": matched, "items": items,
           **({"note": f"{matched}건 중 {limit}건만 (limit 로 조정)"} if matched > limit else {})}
    if len(hmap) == 1 and max_col > 1:
        # 헤더 오인 의심(실험 10 곁가지 — 가이드 함정 5): 병합 제목이 유일 헤더로 잡히면
        # A열만 읽히는데 success 로 그럴듯하게 나간다. 의심 단서를 응답에 싣는다.
        res["hint"] = (f"⚠️ 헤더가 1열뿐인데 시트는 {max_col}열입니다 — 병합 제목이 헤더로 "
                       f"잡혔을 수 있습니다. header_row 파라미터로 실제 헤더 행을 지정하세요.")
    return res


def _items_from_prev(tool_input):
    """파이프 결합: items 미지정 시 직전 step 결과에서 items 또는 table{columns,rows}를 취한다."""
    import json as _json
    prev = tool_input.get("_prev_result") or ""
    if not prev:
        return None
    try:
        data = _json.loads(prev) if isinstance(prev, str) else prev
    except (ValueError, TypeError):
        return None
    if not isinstance(data, dict):
        return None
    if isinstance(data.get("items"), list) and data["items"]:
        return [it for it in data["items"] if isinstance(it, dict)]
    tbl = data.get("table")
    if isinstance(tbl, dict) and tbl.get("columns") and isinstance(tbl.get("rows"), list):
        cols = [str(c) for c in tbl["columns"]]
        return [dict(zip(cols, row)) for row in tbl["rows"]]
    return None


def op_append(tool_input):
    """마지막 데이터 행 뒤에 행 추가 — 기존 수식·서식 불변. 값이 "=..." 문자열이면 수식으로 들어간다."""
    items = tool_input.get("items")
    if isinstance(items, dict) and not isinstance(items.get("items"), list) \
            and not (isinstance(items.get("columns"), list) and isinstance(items.get("rows"), list)) \
            and not isinstance(items.get("table"), dict):
        items = [items]                       # 행 하나를 직접 준 경우
    else:
        # `items: "$변수"`(JSON 문자열·통화 봉투·columns/rows) — 되읽기는 몸의 정본 하나 (B53-2 census)
        try:
            from common.currency import coerce_items_payload
            _rows = coerce_items_payload(items)
            if _rows is not None:
                items = _rows
        except ImportError:
            pass
    if not items:
        items = _items_from_prev(tool_input)
    if not items or not isinstance(items, list):
        return {"success": False,
                "error": "items 가 필요합니다 — [{열이름: 값}, …] 목록 (또는 파이프 직전 step 의 items/table 자동 사용)."}
    items = [it for it in items if isinstance(it, dict) and it]
    if not items:
        return {"success": False, "error": "items 에 유효한 행({열이름: 값})이 없습니다."}

    ctx, err = _prep(tool_input)
    if err:
        return err
    p, wb, ws, hmap = ctx["p"], ctx["wb"], ctx["ws"], ctx["hmap"]
    err = _guard(tool_input, p)
    if err:
        wb.close()
        return err
    all_keys = {k for it in items for k in it if not str(k).startswith("_")}
    err = _validate_keys(all_keys, hmap, "items")
    if err:
        wb.close()
        return err

    last = _last_data_row(ws, hmap, ctx["header_row"])
    start = last + 1
    new_hi = start + len(items) - 1
    # 쓰기 전에 판정(새로 쓴 셀의 수식이 스캔에 섞이지 않게) — 보강④ + 라벨 보조 축
    flags, scan_truncated = _aggregates_missing_rows(wb, ws, ctx["header_row"], start, new_hi)
    label_row = _totals_label_row(ws, hmap, last, ctx["header_row"])
    for i, it in enumerate(items):
        for k, v in it.items():
            if str(k).startswith("_"):
                continue
            cell = ws.cell(row=start + i, column=hmap[k], value=v)
            # 열 표시 형식 승계(⑲ 곁가지) — 위 줄은 '#,##0"원"' 인데 새 줄만 General 이던 것
            ref = ws.cell(row=start + i - 1, column=hmap[k]) if start + i - 1 > ctx["header_row"] else None
            if ref is not None and ref.number_format and ref.number_format != "General":
                cell.number_format = ref.number_format
    _save(wb, p)
    wb.close()
    res = {"success": True, "sheet": ws.title, "appended": len(items),
           "rows": f"{start}~{new_hi}", "path": str(p)}
    if flags or label_row:
        # ⑲ 정직층: 파일은 멀쩡하고 숫자만 틀리는 부류 — 최소한 침묵하지 않는다
        res["totals_row_suspected"] = (flags[0]["row"] if flags and flags[0]["sheet"] == ws.title
                                       else (label_row or (flags[0]["row"] if flags else None)))
        if flags:
            res["aggregates_missing_new_rows"] = [
                {"cell": f"{f['sheet']}!{f['cell']}", "formula": f["formula"]} for f in flags]
        shown = "; ".join(f"{f['sheet']}!{f['cell']}={f['formula']}" for f in flags[:3]) \
            or f"{ws.title} {label_row}행 합계 라벨"
        res["warning"] = (f"⚠️ 새 행({res['rows']})을 참조하지 않는 집계가 있습니다: {shown} — "
                          f"합계/소계가 새 행을 세지 않습니다. 수식 범위를 갱신하거나 행 배치를 확인하세요.")
    if scan_truncated:
        # ⑳-1(실험 12): 상한에 걸려 못 본 구간의 침묵 금지 — 큰 장부에서 무경고 통과가
        # "확인됨"으로 읽히면 안 된다(F9 실측: 3,100행 장부의 바닥 합계를 라벨 없이는 놓침)
        res["scan_truncated"] = True
        note = (f"⚠️ 수식 스캔이 상한(같은 시트 {_SCAN_MAX_SAME}행·교차 시트 {_SCAN_MAX_OTHER}행)에서 "
                f"잘렸습니다 — 상한 밖의 합계·집계는 확인하지 못했습니다.")
        res["warning"] = (res.get("warning", "") + " " + note).strip()
    return res


def op_update(tool_input):
    """where 로 찾은 행들의 set({열: 값}) 셀 수정 — where 필수(전행 갱신 방지)."""
    set_map = tool_input.get("set") or {}
    if not isinstance(set_map, dict) or not set_map:
        return {"success": False, "error": "set 이 필요합니다 — {열이름: 새값}."}

    ctx, err = _prep(tool_input, need_where=True)
    if err:
        return err
    p, wb, ws, hmap = ctx["p"], ctx["wb"], ctx["ws"], ctx["hmap"]
    err = _guard(tool_input, p)
    if err:
        wb.close()
        return err
    err = _validate_keys(set_map.keys(), hmap, "set")
    if err:
        wb.close()
        return err

    rows = list(_matched_rows(ctx))
    if not rows:
        wb.close()
        return {"success": False, "matched": 0,
                "error": f"일치하는 행이 없습니다 — op:find 로 where 값을 먼저 확인하세요. 이 시트의 열: {', '.join(hmap)}"}
    replaced_formulas = {}
    for r in rows:
        for k, v in set_map.items():
            cell = ws.cell(row=r, column=hmap[k])
            # ⑳(2026-08-08 실험 10): 살아있는 수식을 리터럴로 덮을 때 침묵하지 않는다 —
            # 원 수식을 응답에 실어야 되돌릴 근거가 남는다(=SUM 합계를 999 로 덮는 사고 실측).
            if isinstance(cell.value, str) and cell.value.startswith("="):
                replaced_formulas[cell.coordinate] = cell.value
            cell.value = v
    _save(wb, p)
    wb.close()
    res = {"success": True, "sheet": ws.title, "matched": len(rows),
           "updated_rows": rows[:20], "set": {k: _jsonable(v) for k, v in set_map.items()},
           "path": str(p)}
    if replaced_formulas:
        res["replaced_formulas"] = replaced_formulas
        res["warning"] = (f"⚠️ 살아있는 수식 {len(replaced_formulas)}개를 리터럴 값으로 덮었습니다 — "
                          f"원 수식: {'; '.join(f'{c}: {f}' for c, f in list(replaced_formulas.items())[:3])}. "
                          f"의도가 아니면 op:update 로 원 수식을 되살리세요.")
    return res
